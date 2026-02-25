"""
Script para subir empleados desde Empleados.xlsx a Supabase.

Tablas destino:
  1. empleados           → datos principales del Excel
  2. empleados_sst       → fila vacía ligada por empleado_id
  3. empleados_talento_humano → fila vacía ligada por empleado_id

Uso: python upload_empleados.py
"""

import openpyxl
import requests
import json
from datetime import datetime, date

# ─────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────
SUPABASE_URL = "https://ulboklgzjriatmaxzpsi.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InVsYm9rbGd6anJpYXRtYXh6cHNpIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NjI3ODQxNDAsImV4cCI6MjA3ODM2MDE0MH0.gY6_K4JQoJxPZmdXMIbFZfiJAOdavbg8jDJW1rOUSPk"
EXCEL_PATH   = "Empleados.xlsx"

HEADERS = {
    "apikey":        SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type":  "application/json",
    "Prefer":        "return=representation",
}

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def clean(value):
    """Convierte valores Excel a tipos seguros para JSON."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        v = value.strip()
        return v if v else None
    if isinstance(value, bool):
        return value
    return value


def post(table, payload):
    """Inserta una fila en Supabase y devuelve la fila creada."""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    resp = requests.post(url, headers=HEADERS, data=json.dumps(payload))
    if resp.status_code not in (200, 201):
        raise RuntimeError(f"[{table}] HTTP {resp.status_code}: {resp.text}")
    return resp.json()


# ─────────────────────────────────────────────
# LECTURA DEL EXCEL
# ─────────────────────────────────────────────
print("📂 Leyendo Empleados.xlsx …")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

headers = [cell.value for cell in ws[1]]
print(f"   Columnas: {headers}")
print(f"   Filas de datos: {ws.max_row - 1}\n")

# ─────────────────────────────────────────────
# PROCESAMIENTO FILA POR FILA
# ─────────────────────────────────────────────
ok = 0
skipped = 0
errors = []

for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    row_data = dict(zip(headers, row))

    # Saltar filas completamente vacías
    if all(v is None for v in row_data.values()):
        skipped += 1
        continue

    doc = clean(row_data.get("documento_identidad"))
    nombre = clean(row_data.get("nombres"))
    apellido = clean(row_data.get("apellidos"))

    # Saltar filas sin identificador mínimo
    if not doc and not nombre:
        skipped += 1
        continue

    # ── 1. Insertar en `empleados` ──────────────────────────────────
    empleado_payload = {
        "activo":               clean(row_data.get("activo")) if clean(row_data.get("activo")) is not None else True,
        "fecha_ingreso":        clean(row_data.get("fecha_ingreso")),
        "cargo":                clean(row_data.get("cargo")),
        "tipo_vinculacion":     clean(row_data.get("tipo_vinculacion")),
        "apellidos":            apellido,
        "nombres":              nombre,
        "documento_identidad":  doc,
        "rh":                   clean(row_data.get("RH")),
        "correo":               clean(row_data.get("correo")),
        "telefono":             clean(row_data.get("telefono")),
        "direccion":            clean(row_data.get("direccion")),
        "tipo_empleado":        clean(row_data.get("tipo_empleado")),
        "fecha_retiro":         clean(row_data.get("fecha_retiro")),
        "motivo_retiro":        clean(row_data.get("motivo_retiro")),
    }
    # Eliminar claves con valor None para no pisar defaults de la BD
    empleado_payload = {k: v for k, v in empleado_payload.items() if v is not None}

    try:
        result = post("empleados", empleado_payload)
        # Supabase devuelve lista cuando Prefer=return=representation
        empleado_id = result[0]["id"] if isinstance(result, list) else result["id"]

        # ── 2. Fila vacía en `empleados_sst` ────────────────────────
        post("empleados_sst", {"empleado_id": empleado_id})

        # ── 3. Fila vacía en `empleados_talento_humano` ─────────────
        post("empleados_talento_humano", {"empleado_id": empleado_id})

        ok += 1
        print(f"   ✅ [{row_idx}] {apellido}, {nombre} (doc: {doc}) → id {empleado_id}")

    except RuntimeError as err:
        errors.append({"fila": row_idx, "nombre": f"{apellido}, {nombre}", "error": str(err)})
        print(f"   ❌ [{row_idx}] {apellido}, {nombre} → {err}")
    except Exception as err:
        errors.append({"fila": row_idx, "nombre": f"{apellido}, {nombre}", "error": str(err)})
        print(f"   ❌ [{row_idx}] {apellido}, {nombre} → Error inesperado: {err}")

# ─────────────────────────────────────────────
# RESUMEN
# ─────────────────────────────────────────────
print("\n" + "="*50)
print(f"✅ Insertados exitosamente : {ok}")
print(f"⚠️  Filas omitidas (vacías): {skipped}")
print(f"❌ Errores                 : {len(errors)}")

if errors:
    print("\nDetalle de errores:")
    for e in errors:
        print(f"  Fila {e['fila']} | {e['nombre']} → {e['error']}")
