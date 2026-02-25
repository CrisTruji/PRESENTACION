"""Genera insert_empleados.sql desde Empleados.xlsx"""
import openpyxl
from datetime import datetime, date

def esc(v):
    if v is None:
        return 'NULL'
    if isinstance(v, bool):
        return 'TRUE' if v else 'FALSE'
    if isinstance(v, (datetime, date)):
        return f"'{v.strftime('%Y-%m-%d')}'"
    s = str(v).strip()
    if not s:
        return 'NULL'
    s = s.replace("'", "''")
    return f"'{s}'"

wb = openpyxl.load_workbook('Empleados.xlsx')
ws = wb.active
headers = [cell.value for cell in ws[1]]

seen_docs = set()
rows = []
dupes = []
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    d = dict(zip(headers, row))
    if all(v is None for v in d.values()):
        continue
    doc = str(d.get('documento_identidad', '') or '').strip()
    if doc and doc in seen_docs:
        dupes.append((i, doc, str(d.get('nombres','')).strip(), str(d.get('apellidos','')).strip()))
        continue
    if doc:
        seen_docs.add(doc)
    rows.append(d)

lines = []
lines.append("-- =======================================================")
lines.append(f"-- MIGRACION EMPLEADOS - {len(rows)} registros")
lines.append(f"-- Duplicados omitidos: {len(dupes)}")
for d in dupes:
    lines.append(f"--   Fila {d[0]}: doc={d[1]} ({d[2]} {d[3]})")
lines.append("-- =======================================================")
lines.append("")
lines.append("BEGIN;")
lines.append("")
lines.append("-- -------------------------------------------------------")
lines.append("-- 1. TABLA empleados")
lines.append("-- -------------------------------------------------------")

for d in rows:
    activo  = esc(d.get('activo') if d.get('activo') is not None else True)
    fi      = esc(d.get('fecha_ingreso'))
    cargo   = esc(d.get('cargo'))
    tv      = esc(d.get('tipo_vinculacion'))
    ap      = esc(d.get('apellidos'))
    nom     = esc(d.get('nombres'))
    doc     = esc(d.get('documento_identidad'))
    rh      = esc(d.get('RH'))
    correo  = esc(d.get('correo'))
    tel     = esc(d.get('telefono'))
    dir_    = esc(d.get('direccion'))
    te      = esc(d.get('tipo_empleado'))
    fr      = esc(d.get('fecha_retiro'))
    mr      = esc(d.get('motivo_retiro'))

    col_list = ["activo","fecha_ingreso","cargo","tipo_vinculacion",
                "apellidos","nombres","documento_identidad",
                "correo","telefono","direccion","tipo_empleado",
                "fecha_retiro","motivo_retiro"]
    val_list = [activo, fi, cargo, tv, ap, nom, doc, correo, tel, dir_, te, fr, mr]

    if rh != 'NULL':
        col_list.append("rh")
        val_list.append(rh)

    cols = ", ".join(col_list)
    vals = ", ".join(val_list)
    lines.append(f"INSERT INTO empleados ({cols}) VALUES ({vals});")

lines.append("")
lines.append("-- -------------------------------------------------------")
lines.append("-- 2. TABLA empleados_sst (fila vinculada por empleado_id)")
lines.append("-- -------------------------------------------------------")
lines.append("INSERT INTO empleados_sst (empleado_id)")
lines.append("SELECT id FROM empleados")
lines.append("WHERE id NOT IN (SELECT empleado_id FROM empleados_sst WHERE empleado_id IS NOT NULL);")

lines.append("")
lines.append("-- -------------------------------------------------------")
lines.append("-- 3. TABLA empleados_talento_humano (fila vinculada)")
lines.append("-- -------------------------------------------------------")
lines.append("INSERT INTO empleados_talento_humano (empleado_id)")
lines.append("SELECT id FROM empleados")
lines.append("WHERE id NOT IN (SELECT empleado_id FROM empleados_talento_humano WHERE empleado_id IS NOT NULL);")

lines.append("")
lines.append("COMMIT;")
lines.append("")
lines.append("-- Verificacion final:")
lines.append("SELECT COUNT(*) AS empleados FROM empleados;")
lines.append("SELECT COUNT(*) AS sst FROM empleados_sst;")
lines.append("SELECT COUNT(*) AS talento_humano FROM empleados_talento_humano;")

sql = "\n".join(lines)

with open("insert_empleados.sql", "w", encoding="utf-8") as f:
    f.write(sql)

print(f"Archivo generado: insert_empleados.sql")
print(f"Empleados a insertar : {len(rows)}")
print(f"Duplicados omitidos  : {len(dupes)}")
print(f"Lineas totales       : {len(lines)}")
