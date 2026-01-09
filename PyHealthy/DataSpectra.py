import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import pandas as pd
import os
import subprocess
import time
import pyautogui
from supabase import create_client, Client
from typing import Optional
import customtkinter as ctk
import threading
import math
from datetime import datetime
import shutil
import requests
import ctypes
from ctypes import wintypes
import glob
import string
import re
from tkcalendar import DateEntry
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import json
import traceback
from supabase.lib.client_options import ClientOptions

# === CONFIGURACIÓN ===
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CARPETA_BOTONES = os.path.join(BASE_DIR, "Buttons")
URL_TECFOOD = "https://food.teknisa.com//df/#/df_entrada#dfe11000_lancamento_entrada"
URL_RETIRADA = "https://food.teknisa.com//est/#/est_relatorios#est31100_posicao_de_estoque"

SUPABASE_URL = "https://ulboklgzjriatmaxzpsi.supabase.co"
SUPABASE_KEY = "sb_secret_R-yFmIRlPNITmp86n5Ix-g_4Tvf3qAk"
# Variable global para la sesión
session = None
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# === CONFIGURACIÓN DE AUTOMATIZACIÓN ===
from pathlib import Path
import requests

TEMP_PDF_DIR = Path("PDF_Descargados")
TEMP_PDF_DIR.mkdir(exist_ok=True)

# Estados de procesamiento
ESTADO_PENDIENTE = "pendiente"
ESTADO_PROCESANDO = "procesando"
ESTADO_SUBIENDO_TECFOOD = "subiendo_tecfood"
ESTADO_COMPLETADO = "completado"
ESTADO_ERROR = "error"

MAX_INTENTOS = 3

# === VARIABLES GLOBALES ===
archivo_excel = None
df_facturas = pd.DataFrame()
productos_por_factura = {}
codigo_clinica = ""
origen_datos = ""
archivo_retirada = None
df_retirada = pd.DataFrame()
unidad_cargada = ""
carpeta_inventario_actual = ""
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")
fecha_personalizada = None

# PALETA DE COLORES
PRIMARY = "#3A7BD5"
PRIMARY_HOVER = "#5EA0FF"
ACCENT = "#FFA726" 
ACCENT_HOVER = "#FFB74D"
CARD_BG = "#1C1E23"
TEXT_MAIN = "#FFFFFF"
TEXT_SECOND = "#B0B0B0"
BG = "#0E0F12"

def silenciar_customtkinter():
    import sys
    class NullWriter:
        def write(self, txt):
            pass
        def flush(self):
            pass
    sys.stderr = NullWriter()

# ==================== SISTEMA DE LOGIN ====================
class LoginApp:
    def __init__(self, master):
        self.master = master
        self.master.title("DataSpectra - Login")
        self.master.geometry("420x520")
        self.master.configure(fg_color=BG)
        self.master.resizable(False, False)
        
        # Centrar ventana
        self.center_window(420, 520)
        
        # Frame principal del login
        self.login_frame = ctk.CTkFrame(self.master, fg_color=CARD_BG, corner_radius=20)
        self.login_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Título
        self.title_label = ctk.CTkLabel(
            self.login_frame,
            text="🔐 DataSpectra",
            font=("Segoe UI", 26, "bold"),
            text_color=TEXT_MAIN
        )
        self.title_label.pack(pady=(40, 10))
        
        self.subtitle_label = ctk.CTkLabel(
            self.login_frame,
            text="Inicio de Sesión",
            font=("Segoe UI", 14),
            text_color=TEXT_SECOND
        )
        self.subtitle_label.pack(pady=(0, 30))
        
        # Formulario
        self.email_label = ctk.CTkLabel(
            self.login_frame,
            text="Correo Electrónico",
            font=("Segoe UI", 12),
            text_color=TEXT_MAIN,
            anchor="w"
        )
        self.email_label.pack(fill="x", padx=40, pady=(10, 5))
        
        self.email_entry = ctk.CTkEntry(
            self.login_frame,
            placeholder_text="usuario@ejemplo.com",
            width=300,
            height=45,
            font=("Segoe UI", 13),
            fg_color="#252833",
            border_color=PRIMARY,
            corner_radius=12
        )
        self.email_entry.pack(pady=(0, 15))
        
        self.password_label = ctk.CTkLabel(
            self.login_frame,
            text="Contraseña",
            font=("Segoe UI", 12),
            text_color=TEXT_MAIN,
            anchor="w"
        )
        self.password_label.pack(fill="x", padx=40, pady=(5, 5))
        
        self.password_entry = ctk.CTkEntry(
            self.login_frame,
            placeholder_text="••••••••",
            width=300,
            height=45,
            font=("Segoe UI", 13),
            fg_color="#252833",
            border_color=PRIMARY,
            corner_radius=12,
            show="•"
        )
        self.password_entry.pack(pady=(0, 10))
        self.password_entry.bind("<Return>", lambda e: self.login())
        
        # Botón de login
        self.login_button = ctk.CTkButton(
            self.login_frame,
            text="Iniciar Sesión",
            command=self.login,
            width=300,
            height=48,
            font=("Segoe UI", 15, "bold"),
            fg_color=PRIMARY,
            hover_color=PRIMARY_HOVER,
            corner_radius=12
        )
        self.login_button.pack(pady=(20, 10))
        
        # Estado
        self.status_label = ctk.CTkLabel(
            self.login_frame,
            text="",
            font=("Segoe UI", 12),
            text_color=TEXT_SECOND
        )
        self.status_label.pack(pady=(5, 20))
        
        # Footer
        self.footer_label = ctk.CTkLabel(
            self.login_frame,
            text="Solo usuarios registrados en Supabase",
            font=("Segoe UI", 11),
            text_color="#666"
        )
        self.footer_label.pack(pady=(10, 20))
        
        # Intentar auto-login si hay sesión guardada
        self.attempt_auto_login()
    
    def center_window(self, width, height):
        """Centra la ventana en la pantalla"""
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        self.master.geometry(f"{width}x{height}+{x}+{y}")
    
    def attempt_auto_login(self):
        """Intenta auto-login con sesión guardada"""
        try:
            if os.path.exists("session.json"):
                with open("session.json", "r") as f:
                    session_data = json.load(f)
                    if "access_token" in session_data:
                        # Verificar si el token aún es válido
                        global supabase
                        try:
                            user = supabase.auth.get_user(session_data["access_token"])
                            if user:
                                self.status_label.configure(text="Sesión recuperada...", text_color=ACCENT)
                                self.master.after(1000, lambda: self.on_login_success(session_data))
                                return True
                        except:
                            pass
        except Exception as e:
            print(f"Error en auto-login: {e}")
        return False
    
    def save_session(self, session_data):
        """Guarda la sesión en un archivo local"""
        try:
            with open("session.json", "w") as f:
                json.dump({
                    "access_token": session_data.access_token,
                    "refresh_token": session_data.refresh_token,
                    "user_id": session_data.user.id,
                    "email": session_data.user.email
                }, f)
        except Exception as e:
            print(f"Error guardando sesión: {e}")
    
    def clear_session(self):
        """Elimina la sesión guardada"""
        try:
            if os.path.exists("session.json"):
                os.remove("session.json")
        except:
            pass
    
    def login(self):
        """Maneja el proceso de login"""
        email = self.email_entry.get().strip()
        password = self.password_entry.get()
        
        if not email or not password:
            self.status_label.configure(text="Por favor completa todos los campos", text_color="#FF5252")
            return
        
        # Deshabilitar botón durante el login
        self.login_button.configure(state="disabled", text="Iniciando sesión...")
        self.status_label.configure(text="Verificando credenciales...", text_color=ACCENT)
        
        # Ejecutar en un hilo separado
        threading.Thread(target=self.perform_login, args=(email, password), daemon=True).start()
    
    def perform_login(self, email, password):
        """Realiza el login con Supabase"""
        try:
            global supabase, session
            
            # Limpiar cualquier sesión previa
            supabase.auth.sign_out()
            
            # Intentar login
            auth_response = supabase.auth.sign_in_with_password({
                "email": email,
                "password": password
            })
            
            session = auth_response.session
            user = auth_response.user
            
            # Guardar sesión
            self.save_session(session)
            
            # Actualizar interfaz desde el hilo principal
            self.master.after(0, lambda: self.on_login_success(session))
            
        except Exception as e:
            error_msg = str(e)
            if "Invalid login credentials" in error_msg:
                error_msg = "Credenciales inválidas"
            elif "Email not confirmed" in error_msg:
                error_msg = "Email no confirmado"
            elif "rate limit" in error_msg.lower():
                error_msg = "Demasiados intentos. Intenta más tarde"
            
            self.master.after(0, lambda: self.on_login_failure(error_msg))
    
    def on_login_success(self, session_data):
        """Maneja el éxito del login"""
        self.status_label.configure(text="✓ Sesión iniciada correctamente", text_color="#4CAF50")
        
        # Configurar cliente autenticado
        global supabase
        supabase = create_client(
            SUPABASE_URL, 
            SUPABASE_KEY,
            options=ClientOptions(
                auto_refresh_token=True,
                persist_session=True
            )
        )
        
        # Configurar sesión
        if hasattr(session_data, 'access_token'):
            supabase.auth.set_session(session_data.access_token, session_data.refresh_token)
        
        # Esperar un momento y cerrar login
        self.master.after(1500, self.close_login)
    
    def on_login_failure(self, error_msg):
        """Maneja el fallo del login"""
        self.login_button.configure(state="normal", text="Iniciar Sesión")
        self.status_label.configure(text=f"✗ Error: {error_msg}", text_color="#FF5252")
        self.clear_session()  # Limpiar sesión inválida
    
    def close_login(self):
        """Cierra la ventana de login"""
        self.master.destroy()

def iniciar_login():
    """Inicia la ventana de login"""
    login_root = ctk.CTk()
    app = LoginApp(login_root)
    login_root.mainloop()
    
    # Verificar si el login fue exitoso
    try:
        global supabase
        user = supabase.auth.get_user()
        return user is not None
    except:
        return False

# SISTEMA DE NOTIFICACIONES OPTIMIZADO - CORREGIDO
_toast_lock = threading.Lock()
_active_toasts = []

def _position_toast_window(win, width=380, height=96, margin_x=24, margin_y=24, offset_index=0):
    try:
        root_x = root.winfo_rootx()
        root_y = root.winfo_y()
        root_w = root.winfo_width()
        root_h = root.winfo_height()
    except Exception:
        root_x, root_y, root_w, root_h = 0, 0, win.winfo_screenwidth(), win.winfo_screenheight()

    x = root_x + max(root_w - width - margin_x, 0)
    y = root_y + max(root_h - height - margin_y - offset_index * (height + 12), 0)
    win.geometry(f"{width}x{height}+{x}+{y}")

def _reposition_toasts():
    with _toast_lock:
        for idx, t in enumerate(list(_active_toasts)):
            try:
                if t.winfo_exists():
                    _position_toast_window(t, width=t.winfo_width() or 360, height=t.winfo_height() or 100, offset_index=idx)
            except Exception:
                pass

def mostrar_toast(mensaje, tipo="info", duracion=3200, titulo=None):
    estilos = {
        "info":    {"icon": "ℹ️", "bg": "#0b1220", "dot": "#3A7BD5", "border": "#1E3A8A", "shadow": "#1E40AF"},
        "success": {"icon": "✅", "bg": "#07170b", "dot": "#16A34A", "border": "#166534", "shadow": "#15803D"},
        "warning": {"icon": "⚠️", "bg": "#1f1200", "dot": "#F59E0B", "border": "#92400E", "shadow": "#B45309"},
        "error":   {"icon": "❌", "bg": "#2b0e0e", "dot": "#EF4444", "border": "#991B1B", "shadow": "#DC2626"},
    }
    st = estilos.get(tipo, estilos["info"])

    TOAST_W = 360
    TOAST_H = 100

    with _toast_lock:
        offset_index = len(_active_toasts)

    toast = ctk.CTkToplevel(root)
    toast.overrideredirect(True)
    toast.attributes("-topmost", True)
    toast.configure(fg_color=st["bg"])
    
    # Frame principal
    main_frame = ctk.CTkFrame(toast, fg_color=st["bg"], corner_radius=16)
    main_frame.pack(fill="both", expand=True, padx=1, pady=1)
    
    # Frame de sombra
    shadow_frame = ctk.CTkFrame(main_frame, fg_color=st["shadow"], height=4, corner_radius=8)
    shadow_frame.pack(side="bottom", fill="x", padx=8, pady=(0, 2))
    shadow_frame.pack_propagate(False)
    
    # Frame interno
    wrapper = ctk.CTkFrame(main_frame, fg_color=st["bg"], corner_radius=14, width=TOAST_W-12, height=TOAST_H-12)
    wrapper.pack_propagate(False)
    wrapper.pack(fill="both", expand=False, padx=6, pady=(6, 4))

    # Frame del icono
    icon_frame = ctk.CTkFrame(wrapper, fg_color=st["bg"], width=56, height=56)
    icon_frame.pack_propagate(False)
    icon_frame.pack(side="left", padx=(12, 16), pady=12)

    # Canvas para el círculo
    dot_size = 48
    dot = tk.Canvas(icon_frame, width=56, height=56, highlightthickness=0, bg=st["bg"])
    
    circle_x1 = (56 - dot_size) / 2
    circle_y1 = (56 - dot_size) / 2
    circle_x2 = circle_x1 + dot_size
    circle_y2 = circle_y1 + dot_size
    
    dot.create_oval(circle_x1, circle_y1, circle_x2, circle_y2, fill=st["dot"], outline=st["dot"], width=0)
    
    text_x = 56 / 2
    text_y = 56 / 2
    
    dot.create_text(text_x, text_y, text=st["icon"], font=("Segoe UI Emoji", 16), fill="#FFFFFF")
    dot.pack(fill="both", expand=True)

    # Contenedor de texto
    text_container = ctk.CTkFrame(wrapper, fg_color=st["bg"], corner_radius=0)
    text_container.pack(side="left", fill="both", expand=True, pady=12, padx=(0, 12))

    if titulo:
        lbl_title = ctk.CTkLabel(text_container, text=titulo, font=("Segoe UI", 12, "bold"), text_color="#FFFFFF", anchor="w")
        lbl_title.pack(fill="x", pady=(0, 2))
        lbl_msg = ctk.CTkLabel(text_container, text=mensaje, font=("Segoe UI", 11), text_color="#C7CCD1", wraplength=TOAST_W-140, anchor="w", justify="left")
        lbl_msg.pack(fill="both")
    else:
        lbl_msg = ctk.CTkLabel(text_container, text=mensaje, font=("Segoe UI", 12), text_color="#E6E9EE", wraplength=TOAST_W-140, anchor="w", justify="left")
        lbl_msg.pack(fill="both")

    def close_toast(e=None):
        if toast.winfo_exists():
            with _toast_lock:
                if toast in _active_toasts:
                    _active_toasts.remove(toast)
            try:
                toast.destroy()
            except:
                pass
            _reposition_toasts()
    
    # Bind click para cerrar
    for widget in [wrapper, text_container, lbl_msg, icon_frame, dot, main_frame, shadow_frame]:
        try:
            widget.bind("<Button-1>", close_toast)
        except:
            pass
    if titulo:
        try:
            lbl_title.bind("<Button-1>", close_toast)
        except:
            pass

    toast.update_idletasks()
    _position_toast_window(toast, width=TOAST_W, height=TOAST_H, offset_index=offset_index)
    
    try:
        toast.attributes("-alpha", 0.0)
    except:
        pass

    with _toast_lock:
        _active_toasts.append(toast)

    def _animate_in():
        try:
            for i in range(6):
                alpha = (i + 1) / 6
                try:
                    toast.attributes("-alpha", alpha)
                except:
                    pass
                time.sleep(0.02)
        except:
            pass

    def _animate_out_and_destroy():
        time.sleep(duracion / 1000.0)
        try:
            for i in range(5):
                alpha = max(0.0, 1 - (i + 1) / 5)
                try:
                    toast.attributes("-alpha", alpha)
                except:
                    pass
                time.sleep(0.03)
        except:
            pass
        close_toast()

    threading.Thread(target=_animate_in, daemon=True).start()
    threading.Thread(target=_animate_out_and_destroy, daemon=True).start()

# ANIMACIONES OPTIMIZADAS
def simple_button_hover(button, is_enter=True):
    if is_enter:
        if button == btn_excel:
            button.configure(fg_color=ACCENT_HOVER)
        elif button == btn_supabase:
            button.configure(fg_color=PRIMARY_HOVER)
        elif button == btn_iniciar and button.cget("state") == "normal":
            button.configure(fg_color=PRIMARY_HOVER)
        elif button == btn_volver:
            button.configure(fg_color="#7AB8FF")
        elif button == btn_close or button == btn_exit_menu:
            button.configure(fg_color="#FF6B6B")
        elif button == btn_descargar_informes:
            button.configure(fg_color=ACCENT_HOVER)
        elif button == btn_procesar_informes:
            button.configure(fg_color=PRIMARY_HOVER)
    else:
        if button == btn_excel:
            button.configure(fg_color=ACCENT)
        elif button == btn_supabase:
            button.configure(fg_color=PRIMARY)
        elif button == btn_iniciar and button.cget("state") == "normal":
            button.configure(fg_color=PRIMARY)
        elif button == btn_volver:
            button.configure(fg_color="#5EA0FF")
        elif button == btn_close or button == btn_exit_menu:
            button.configure(fg_color="#FF5252")
        elif button == btn_descargar_informes:
            button.configure(fg_color=ACCENT)
        elif button == btn_procesar_informes:
            button.configure(fg_color=PRIMARY)

def simple_item_hover(item, is_enter=True):
    if is_enter:
        item.configure(fg_color="#1E2025")
    else:
        item.configure(fg_color="#141518")

def quick_pulse_animation(widget, pulse_color, duration=0.3):
    def _pulse():
        try:
            original_color = widget.cget("fg_color")
            widget.configure(fg_color=pulse_color)
            time.sleep(duration)
            widget.configure(fg_color=original_color)
        except Exception:
            pass
    threading.Thread(target=_pulse, daemon=True).start()

def fade_transition(widget, target_alpha, duration=0.2):
    def _fade():
        try:
            steps = 4
            current_alpha = widget.attributes("-alpha") if widget.attributes("-alpha") else 1.0
            
            for i in range(steps):
                progress = (i + 1) / steps
                alpha = current_alpha + (target_alpha - current_alpha) * progress
                widget.attributes("-alpha", alpha)
                time.sleep(duration / steps)
        except Exception:
            widget.attributes("-alpha", target_alpha)
    
    threading.Thread(target=_fade, daemon=True).start()

# Confirmación cerrar app OPTIMIZADA
def confirmar_salida(titulo="Confirmar salida", mensaje="¿Deseas salir de la aplicación?"):
    result = {"value": False}
    width, height = 400, 140

    confirm = ctk.CTkToplevel(root)
    confirm.overrideredirect(True)
    confirm.attributes("-topmost", True)
    confirm.configure(fg_color="#0f1724")
    
    confirm.update_idletasks()
    screen_w = confirm.winfo_screenwidth()
    screen_h = confirm.winfo_screenheight()
    x = (screen_w // 2) - (width // 2)
    y = (screen_h // 2) - (height // 2)
    confirm.geometry(f"{width}x{height}+{x}+{y}")
    
    wrapper = ctk.CTkFrame(confirm, fg_color="#0f1724", corner_radius=14)
    wrapper.pack(fill="both", expand=True, padx=8, pady=8)

    barra = ctk.CTkFrame(wrapper, fg_color="#334155", width=8, corner_radius=14)
    barra.pack(side="left", fill="y", padx=(2, 8), pady=6)

    content = ctk.CTkFrame(wrapper, fg_color="#0f1724", corner_radius=8)
    content.pack(side="left", fill="both", expand=True, padx=(0, 8), pady=8)

    lbl_t = ctk.CTkLabel(content, text=titulo, font=("Segoe UI", 13, "bold"), text_color="#FFFFFF")
    lbl_t.pack(anchor="w", padx=8, pady=(6, 0))

    lbl_m = ctk.CTkLabel(
        content,
        text=mensaje,
        font=("Segoe UI", 11),
        text_color="#9CA3AF",
        wraplength=width - 80,
        anchor="w",
        justify="left"
    )
    lbl_m.pack(anchor="w", padx=8, pady=(4, 8))

    btn_frame = ctk.CTkFrame(content, fg_color="#0f1724")
    btn_frame.pack(anchor="e", padx=8, pady=(0, 8))

    def on_confirm():
        result["value"] = True
        fade_transition(confirm, 0.0, 0.15)
        confirm.after(150, confirm.destroy)

    def on_cancel():
        result["value"] = False
        fade_transition(confirm, 0.0, 0.15)
        confirm.after(150, confirm.destroy)

    btn_cancel = ctk.CTkButton(
        btn_frame,
        text="Cancelar",
        width=110,
        command=on_cancel,
        fg_color="#475569",
        hover_color="#64748B"
    )
    btn_cancel.pack(side="right", padx=(10, 0))

    btn_ok = ctk.CTkButton(
        btn_frame,
        text="Salir",
        width=110,
        command=on_confirm,
        fg_color="#DC2626",
        hover_color="#EF4444"
    )
    btn_ok.pack(side="right")
    
    try:
        confirm.attributes("-alpha", 0.0)
    except Exception:
        pass

    fade_transition(confirm, 1.0, 0.15)
    confirm.wait_window()
    return result["value"]


# ========= FUNCIONES CORREGIDAS PARA INVENTARIO =========

# Lista de códigos de clínica estáticos
CODIGOS_CLINICAS = ["0001", "0011", "0024", "0002", "0031", "0014", "0017", "0018", "0003"]
NOMBRES_CLINICAS = {
    "0001": "HEALTHY MATRIZ",
    "0011": "PLANTA IBAGUE",
    "0024": "SEVIN DRUMMOND",
    "0002": "SEATECH",
    "0031": "CLINICA AZUL MEDPLUS",
    "0014": "BRUNE RETAIL",
    "0017": "CARRITO MEDICADIZ",
    "0018": "CARRITO KERALTY",
    "0003": "ASOCAÑA CALI"
}

def crear_carpeta_inventario(fecha_inventario=None):
    escritorio = os.path.join(os.path.expanduser("~"), "Escritorio")
    if not os.path.isdir(escritorio):
        escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
    
    # Determinar la fecha del inventario
    if fecha_personalizada:
        fecha_inv = fecha_personalizada.replace("/", "-")
    else:
        fecha_inv = datetime.now().strftime("%d-%m-%Y")
    
    # Nombre base de carpeta
    nombre_base = f"Inventario {fecha_inv}"
    fecha_hoy = datetime.now().strftime("%H-%M_%d-%m-%Y")
    
    # Verificar si ya existe la carpeta base
    carpeta_base = os.path.join(escritorio, nombre_base)
    
    if os.path.exists(carpeta_base):
        # Si existe, crear una con sufijo
        nombre_carpeta = f"{nombre_base} Generado a las {fecha_hoy}"
        carpeta_final = os.path.join(escritorio, nombre_carpeta)
        os.makedirs(carpeta_final, exist_ok=True)
        print(f"📁 Carpeta DUPLICADA - Nueva creada: {nombre_carpeta}")
    else:
        # Si no existe, crear la carpeta base
        nombre_carpeta = nombre_base
        carpeta_final = carpeta_base
        os.makedirs(carpeta_final, exist_ok=True)
        print(f"📁 Carpeta NUEVA creada: {nombre_carpeta}")
    
    # Guardar referencia global de carpeta actual
    global carpeta_inventario_actual
    carpeta_inventario_actual = carpeta_final
    
    print(f"📁 Ruta completa: {carpeta_final}")
    return carpeta_final

def obtener_archivo_mas_reciente(carpeta):
    archivos = []
    for nombre in os.listdir(carpeta):
        if nombre.lower().startswith("est31100") and nombre.lower().endswith((".xlsx", ".xls")):
            ruta = os.path.join(carpeta, nombre)
            archivos.append((ruta, os.path.getmtime(ruta)))
    if not archivos:
        return None
    archivos.sort(key=lambda x: x[1], reverse=True)
    return archivos[0][0]

def obtener_carpeta_descargas():
    # ==============================
    # 1. CACHE (si ya se detectó antes)
    # ==============================
    cache_file = "descargas_cache.json"
    try:
        if os.path.exists(cache_file):
            with open(cache_file, "r", encoding="utf8") as f:
                cache = json.load(f)
                ruta_cache = cache.get("descargas")
                if ruta_cache and os.path.isdir(ruta_cache):
                    print(f"📌 Usando ruta guardada (cache): {ruta_cache}")
                    return ruta_cache
    except Exception:
        pass

    # ==============================================
    # 2. RUTA PERSONALIZADA DETECTADA EN TU EMPRESA
    # ==============================================
    rutas_comunes_empresa = [
        r"O:\perfil",
        r"O:\usuarios",
        r"O:\user",
        r"O:\home",
        r"P:\perfil",
        r"P:\usuarios",
        r"S:\perfil",
        r"S:\usuarios"
    ]

    # Buscar patrones típicos corporativos
    for disco in "OPQRSTUVWXYZ":
        raiz = f"{disco}:\\"
        if not os.path.exists(raiz):
            continue

        for root, dirs, files in os.walk(raiz):
            if re.search(r'(Descargas|Downloads)$', root, re.IGNORECASE):
                print(f"🔍 Carpeta de descargas encontrada en red: {root}")
                _guardar_cache_descargas(root)
                return root

    # =============================================================
    # 3. Detectar carpeta redirigida REAL de Windows (shell folders)
    # =============================================================
    try:
        import winreg

        keys = [
            r"SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders",
            r"SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders"
        ]

        for key in keys:
            try:
                reg = winreg.OpenKey(winreg.HKEY_CURRENT_USER, key)
                ruta = winreg.QueryValueEx(reg, "Downloads")[0]

                # Expandir variables de entorno (%USERPROFILE%)
                ruta = os.path.expandvars(ruta)

                if os.path.isdir(ruta):
                    print(f"📌 Detectada carpeta redirigida por Windows: {ruta}")
                    _guardar_cache_descargas(ruta)
                    return ruta

            except Exception:
                continue
    except Exception:
        pass

    # ============================
    # 4. Exploración multi-disco
    # ============================
    nombres = ["Descargas", "Downloads"]

    for disco in "CDEFGHIJKLMNOPQRSTUVWXYZ":
        raiz = f"{disco}:\\"
        if not os.path.exists(raiz):
            continue

        for root, dirs, _ in os.walk(raiz):
            for carpeta in dirs:
                if carpeta in nombres:
                    ruta_final = os.path.join(root, carpeta)
                    print(f"🔍 Encontrada: {ruta_final}")
                    _guardar_cache_descargas(ruta_final)
                    return ruta_final

    # ============================
    # 5. Fallback final
    # ============================
    fallback = os.path.join(os.path.expanduser("~"), "Downloads")
    print(f"⚠️ Usando fallback: {fallback}")
    _guardar_cache_descargas(fallback)
    return fallback

def _guardar_cache_descargas(ruta):
    """ Guarda la ruta válida para acelerar detecciones futuras """
    try:
        with open("descargas_cache.json", "w", encoding="utf8") as f:
            json.dump({"descargas": ruta}, f)
    except Exception:
        pass

def limpiar_carpeta_descargas():
    """Elimina archivos Excel antiguos de inventario de la carpeta de descargas"""
    try:
        carpeta_descargas = obtener_carpeta_descargas()
        archivos_eliminados = 0
        
        for archivo in os.listdir(carpeta_descargas):
            ruta_archivo = os.path.join(carpeta_descargas, archivo)
            # Buscar archivos que contengan EST31100 con cualquier extensión
            if ('EST31100' in archivo or 'Inventario_Clinica_' in archivo):
                try:
                    os.remove(ruta_archivo)
                    archivos_eliminados += 1
                    print(f"🗑️ Eliminado archivo antiguo: {archivo}")
                except Exception as e:
                    print(f"⚠️ No se pudo eliminar {archivo}: {e}")
        
        print(f"🧹 Total de archivos antiguos eliminados: {archivos_eliminados}")
        return archivos_eliminados
        
    except Exception as e:
        print(f"⚠️ Error limpiando carpeta descargas: {e}")
        return 0

def obtener_archivo_mas_reciente(carpeta, patrones=None):
    if patrones is None:
        patrones = ['EST31100*', '*.xlsx', '*.xls']
    archivos = []
    for patron in patrones:
        archivos.extend(glob.glob(os.path.join(carpeta, patron)))
    if not archivos:
        return None
    # ordenar por fecha de modificación (más reciente primero)
    archivos = sorted(archivos, key=os.path.getmtime, reverse=True)
    return archivos[0]

def esperar_archivo_descargado(carpeta_descargas, tiempo_maximo=5, poll_interval=0.8):
    tiempo_inicio = time.time()
    if not carpeta_descargas or not os.path.isdir(carpeta_descargas):
        print(f"Carpeta de descargas inválida: {carpeta_descargas}")
        try:
            mostrar_toast("Carpeta de descargas inválida", tipo="error", titulo="Error")
        except Exception:
            pass
        return None

    print(f"⏳ Esperando archivo en: {carpeta_descargas} (max {tiempo_maximo}s)")
    ultimo_encontrado = None
    while time.time() - tiempo_inicio < tiempo_maximo:
        candidato = obtener_archivo_mas_reciente(carpeta_descargas)
        if candidato and candidato != ultimo_encontrado:
            # comprobar estabilidad y que no sea temporal (.crdownload / .part)
            ruta_final = candidato
            if ruta_final.endswith('.crdownload') or ruta_final.endswith('.part'):
                posible = ruta_final.rsplit('.', 1)[0]
                if os.path.exists(posible):
                    ruta_final = posible
            if os.path.exists(ruta_final) and _archivo_estable(ruta_final):
                print(f"📥 Archivo detectado: {os.path.basename(ruta_final)}")
                try:
                    mostrar_toast(f"Archivo detectado: {os.path.basename(ruta_final)}", tipo="info", titulo="Descarga")
                except Exception:
                    pass
                return ruta_final
            ultimo_encontrado = candidato
        time.sleep(poll_interval)
    print("❌ No se detectó archivo descargado en el tiempo esperado")
    # fallback: buscar EST31100 por si Windows lo listó con retraso
    respaldo = glob.glob(os.path.join(carpeta_descargas, 'EST31100*'))
    if respaldo:
        candidato = max(respaldo, key=os.path.getmtime)
        if _archivo_estable(candidato):
            print(f"⚠️ Usando archivo de respaldo: {os.path.basename(candidato)}")
            return candidato
    return None

def _archivo_estable(ruta):
    """Devuelve True si el archivo existe, tiene tamaño > 0 y puede abrirse para lectura (no bloqueado)."""
    try:
        if not os.path.exists(ruta):
            return False
        tam1 = os.path.getsize(ruta)
        time.sleep(0.8)
        tam2 = os.path.getsize(ruta)
        if tam1 == tam2 and tam1 > 0:
            # Intentar abrir para confirmar que no está bloqueado
            try:
                with open(ruta, 'rb'):
                    pass
                return True
            except Exception:
                return False
        return False
    except Exception:
        return False

def renombrar_y_mover_archivo(archivo_descargado, carpeta_destino, codigo_clinica, indice):
    """
    Renombra el archivo descargado y lo mueve DIRECTAMENTE a la carpeta final.
    """
    try:
        if not archivo_descargado or not os.path.exists(archivo_descargado):
            print(f"❌ Archivo descargado no existe: {archivo_descargado}")
            return None
        
        # Formato de fecha simple
        fecha_actual = datetime.now().strftime("%Y-%m-%d")
        
        # Nombre de la clínica
        nombre_clinica = NOMBRES_CLINICAS.get(codigo_clinica, codigo_clinica)
        
        # Nombre simple del archivo
        nombre_nuevo = f"{nombre_clinica}.xlsx"
        
        ruta_nueva = os.path.join(carpeta_destino, nombre_nuevo)
        
        # Si ya existe, agregar número
        if os.path.exists(ruta_nueva):
            base, ext = os.path.splitext(nombre_nuevo)
            contador = 1
            while os.path.exists(os.path.join(carpeta_destino, f"{base}_{contador}{ext}")):
                contador += 1
            nombre_nuevo = f"{base}_{contador}{ext}"
            ruta_nueva = os.path.join(carpeta_destino, nombre_nuevo)
        
        # Mover directamente a la carpeta final
        shutil.move(archivo_descargado, ruta_nueva)
        
        print(f"✅ Archivo guardado: {nombre_nuevo}")
        return nombre_nuevo
        
    except Exception as e:
        print(f"❌ Error moviendo archivo: {e}")
        return None

def debug_descargas():
    """Función para debug: mostrar qué archivos hay en descargas"""
    carpeta_descargas = obtener_carpeta_descargas()
    print(f"\n🔍 DEBUG - Archivos en descargas:")
    archivos_encontrados = False
    for archivo in os.listdir(carpeta_descargas):
        if 'EST31100' in archivo:
            ruta = os.path.join(carpeta_descargas, archivo)
            tamaño = os.path.getsize(ruta)
            mod_time = time.ctime(os.path.getmtime(ruta))
            print(f"   📄 {archivo} - {tamaño} bytes - {mod_time}")
            archivos_encontrados = True
    
    if not archivos_encontrados:
        print("   ℹ️ No se encontraron archivos EST31100")
    
    # Mostrar todos los archivos Excel también
    print(f"\n🔍 DEBUG - Todos los archivos Excel en descargas:")
    excel_encontrados = False
    for archivo in os.listdir(carpeta_descargas):
        if archivo.endswith(('.xlsx', '.xls', '.síax', '.síac')):
            ruta = os.path.join(carpeta_descargas, archivo)
            tamaño = os.path.getsize(ruta)
            mod_time = time.ctime(os.path.getmtime(ruta))
            print(f"   📊 {archivo} - {tamaño} bytes - {mod_time}")
            excel_encontrados = True
    
    if not excel_encontrados:
        print("   ℹ️ No se encontraron archivos Excel")

# ============================================================
# SISTEMA DE SELECCIÓN DE FECHA PARA INFORMES DE INVENTARIO
# ============================================================

def pedir_fecha_informes():
    global fecha_personalizada
    
    # ===== CREAR VENTANA MODAL CON SCROLL =====
    win = ctk.CTkToplevel(root)
    win.title("Seleccionar fecha para los informes")
    win.geometry("520x420")
    win.resizable(False, False)
    win.configure(fg_color="#0F1724")
    win.attributes("-topmost", True)
    win.grab_set()
    
    # ===== CENTRAR EN PANTALLA =====
    win.update_idletasks()
    W, H = 520, 420
    x = (win.winfo_screenwidth() // 2) - (W // 2)
    y = (win.winfo_screenheight() // 2) - (H // 2)
    win.geometry(f"{W}x{H}+{x}+{y}")
    
    # ===== ANIMACIÓN FADE-IN =====
    def fade_in(alpha=0.0):
        if alpha <= 1.0:
            win.attributes("-alpha", alpha)
            win.after(10, lambda: fade_in(alpha + 0.05))
    fade_in()
    
    # ===== CONTENEDOR PRINCIPAL CON SCROLL =====
    main_container = ctk.CTkScrollableFrame(
        win, 
        fg_color="transparent",
        scrollbar_button_color="#4B5563",
        scrollbar_button_hover_color="#6B7280"
    )
    main_container.pack(fill="both", expand=True, padx=20, pady=20)
    
    # Tarjeta principal
    card = ctk.CTkFrame(main_container, fg_color="#1C1E23", corner_radius=20)
    card.pack(fill="both", expand=True)
    
    # Contenedor interno para contenido fijo
    content_frame = ctk.CTkFrame(card, fg_color="transparent")
    content_frame.pack(fill="both", expand=True, padx=30, pady=25)
    
    # ===== ENCABEZADO =====
    header_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
    header_frame.pack(fill="x", pady=(0, 15))
    
    # Icono e título
    icon_label = ctk.CTkLabel(header_frame, text="📅", font=("Segoe UI Emoji", 32))
    icon_label.pack(side="left", padx=(0, 15))
    
    title_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
    title_frame.pack(side="left", fill="both", expand=True)
    
    ctk.CTkLabel(
        title_frame,
        text="Seleccionar fecha de inventario",
        font=("Segoe UI", 20, "bold"),
        text_color="white"
    ).pack(anchor="w")
    
    ctk.CTkLabel(
        title_frame,
        text="Elige la fecha para la cual deseas generar los informes",
        font=("Segoe UI", 12),
        text_color="#AAB4C0"
    ).pack(anchor="w", pady=(2, 0))
    
    # ===== OPCIONES DE FECHA =====
    options_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
    options_frame.pack(fill="x", pady=(20, 0))
    
    def on_hover_enter(frame, is_hovering):
        if is_hovering:
            frame.configure(fg_color="#2E313C")
        else:
            frame.configure(fg_color="#252833")
    
    # OPCIÓN 1: FECHA ACTUAL
    option1_frame = ctk.CTkFrame(options_frame, fg_color="#252833", corner_radius=12, height=70)
    option1_frame.pack(fill="x", pady=(0, 10))
    option1_frame.pack_propagate(False)
    
    option1_frame.bind("<Enter>", lambda e: on_hover_enter(option1_frame, True))
    option1_frame.bind("<Leave>", lambda e: on_hover_enter(option1_frame, False))
    
    # Contenido opción 1
    option1_content = ctk.CTkFrame(option1_frame, fg_color="transparent")
    option1_content.pack(fill="both", expand=True, padx=20)
    
    option1_text_frame = ctk.CTkFrame(option1_content, fg_color="transparent")
    option1_text_frame.pack(side="left", fill="both", expand=True)
    
    ctk.CTkLabel(
        option1_text_frame,
        text="📆 Usar fecha actual",
        font=("Segoe UI", 15, "bold"),
        text_color="white"
    ).pack(anchor="w")
    
    ctk.CTkLabel(
        option1_text_frame,
        text=f"Hoy: {datetime.now().strftime('%d/%m/%Y')}",
        font=("Segoe UI", 13),
        text_color="#60A5FA"
    ).pack(anchor="w", pady=(5, 0))
    
    # Botón de selección para opción 1
    option1_btn = ctk.CTkButton(
        option1_content,
        text="Seleccionar",
        width=100,
        corner_radius=8,
        fg_color="#3B82F6",
        hover_color="#2563EB",
        font=("Segoe UI", 11, "bold")
    )
    option1_btn.pack(side="right", padx=(10, 0))
    
    # OPCIÓN 2: FECHA PERSONALIZADA
    option2_frame = ctk.CTkFrame(options_frame, fg_color="#252833", corner_radius=12, height=70)
    option2_frame.pack(fill="x")
    option2_frame.pack_propagate(False)
    
    option2_frame.bind("<Enter>", lambda e: on_hover_enter(option2_frame, True))
    option2_frame.bind("<Leave>", lambda e: on_hover_enter(option2_frame, False))
    
    # Contenido opción 2
    option2_content = ctk.CTkFrame(option2_frame, fg_color="transparent")
    option2_content.pack(fill="both", expand=True, padx=20)
    
    option2_text_frame = ctk.CTkFrame(option2_content, fg_color="transparent")
    option2_text_frame.pack(side="left", fill="both", expand=True)
    
    ctk.CTkLabel(
        option2_text_frame,
        text="🗓️ Seleccionar fecha personalizada",
        font=("Segoe UI", 15, "bold"),
        text_color="white"
    ).pack(anchor="w")
    
    ctk.CTkLabel(
        option2_text_frame,
        text="Elige cualquier fecha del calendario",
        font=("Segoe UI", 13),
        text_color="#60A5FA"
    ).pack(anchor="w", pady=(5, 0))
    
    # Botón de selección para opción 2
    option2_btn = ctk.CTkButton(
        option2_content,
        text="Seleccionar",
        width=100,
        corner_radius=8,
        fg_color="#3B82F6",
        hover_color="#2563EB",
        font=("Segoe UI", 11, "bold")
    )
    option2_btn.pack(side="right", padx=(10, 0))
    
    # ===== ÁREA DE FECHA PERSONALIZADA (INICIALMENTE OCULTA) =====
    custom_date_container = ctk.CTkFrame(content_frame, fg_color="transparent")
    custom_date_container.pack(fill="x", pady=(20, 0))
    
    # Frame para el contenido de fecha personalizada
    custom_date_frame = ctk.CTkFrame(custom_date_container, fg_color="#1F2937", corner_radius=12)
    custom_date_frame.pack(fill="x", pady=(5, 0))
    
    # Inicialmente oculto
    custom_date_frame.pack_forget()
    
    def mostrar_selector_fecha():
        # Si ya está visible, no hacer nada
        if custom_date_frame.winfo_ismapped():
            return
            
        # Limpiar contenido previo
        for widget in custom_date_frame.winfo_children():
            widget.destroy()
        
        # Crear contenido de fecha personalizada
        inner_frame = ctk.CTkFrame(custom_date_frame, fg_color="transparent")
        inner_frame.pack(fill="both", expand=True, padx=25, pady=25)
        
        # Título
        ctk.CTkLabel(
            inner_frame,
            text="📆 Selecciona una fecha específica:",
            font=("Segoe UI", 16, "bold"),
            text_color="white"
        ).pack(anchor="w", pady=(0, 15))
        
        # Frame para calendario
        calendar_frame = ctk.CTkFrame(inner_frame, fg_color="#111827", corner_radius=10)
        calendar_frame.pack(fill="x", pady=(0, 20))
        
        # Widget de calendario con estilo mejorado y más grande
        fecha_widget = DateEntry(
            calendar_frame,
            width=24,
            height=8,
            background="#111827",
            foreground="white",
            selectbackground="#3B82F6",
            selectforeground="white",
            normalbackground="#1F2937",
            normalforeground="#D1D5DB",
            weekendbackground="#1F2937",
            weekendforeground="#9CA3AF",
            headersbackground="#111827",
            headersforeground="#60A5FA",
            borderwidth=0,
            relief="flat",
            date_pattern="dd/mm/yyyy",
            font=("Segoe UI", 11),
            justify="center"
        )
        fecha_widget.pack(padx=15, pady=15)
        
        # Frame para botones
        buttons_frame = ctk.CTkFrame(inner_frame, fg_color="transparent")
        buttons_frame.pack(fill="x", pady=(10, 0))
        
        # Botón para confirmar fecha personalizada
        def confirmar_fecha_personalizada():
            selected_date = fecha_widget.get()
            if selected_date:
                globals()["fecha_personalizada"] = selected_date
                mostrar_toast(f"Fecha seleccionada: {selected_date}", tipo="success", titulo="Fecha confirmada")
                fade_out_and_close(win)
                iniciar_descarga()
            else:
                mostrar_toast("Por favor selecciona una fecha", tipo="warning", titulo="Fecha requerida")
        
        # Botón para cancelar selección personalizada
        def cancelar_personalizada():
            ocultar_selector_fecha()
        
        ctk.CTkButton(
            buttons_frame,
            text="Confirmar fecha seleccionada",
            fg_color="#3B82F6",
            hover_color="#2563EB",
            corner_radius=10,
            font=("Segoe UI", 13, "bold"),
            height=42,
            command=confirmar_fecha_personalizada
        ).pack(side="left", padx=(0, 10))
        
        ctk.CTkButton(
            buttons_frame,
            text="Cancelar",
            fg_color="transparent",
            hover_color="#374151",
            border_width=1,
            border_color="#4B5563",
            text_color="#9CA3AF",
            corner_radius=10,
            font=("Segoe UI", 13),
            height=42,
            command=cancelar_personalizada
        ).pack(side="left")
        
        # Mostrar con animación
        custom_date_frame.pack(fill="x", pady=(5, 0))
        
        # Forzar actualización y scroll a la vista
        main_container._parent_canvas.yview_moveto(1.0)
        win.update_idletasks()
    
    def ocultar_selector_fecha():
        custom_date_frame.pack_forget()
        win.update_idletasks()
    
    # ===== FUNCIONES PARA OPCIONES =====
    def usar_fecha_actual():
        globals()["fecha_personalizada"] = None
        mostrar_toast(f"Usando fecha actual: {datetime.now().strftime('%d/%m/%Y')}", tipo="success", titulo="Fecha actual")
        fade_out_and_close(win)
        iniciar_descarga()
    
    # Asignar funciones a los botones
    option1_btn.configure(command=usar_fecha_actual)
    option2_btn.configure(command=mostrar_selector_fecha)
    
    # ===== ANIMACIÓN DE CIERRE =====
    def fade_out_and_close(toplevel, alpha=1.0):
        if alpha > 0:
            toplevel.attributes("-alpha", alpha)
            toplevel.after(8, lambda: fade_out_and_close(toplevel, alpha - 0.1))
        else:
            toplevel.destroy()
    
    # ===== INICIAR DESCARGA =====
    def iniciar_descarga():
        mostrar_pantalla("inventario")
        root.after(300, descargar_informes_inventario)
    
    # ===== BOTÓN DE CANCELAR GENERAL =====
    cancel_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
    cancel_frame.pack(fill="x", pady=(20, 0))
    
    def cancelar():
        mostrar_toast("Selección de fecha cancelada", tipo="info", titulo="Cancelado")
        fade_out_and_close(win)
    
    cancel_btn = ctk.CTkButton(
        cancel_frame,
        text="Cancelar selección",
        fg_color="transparent",
        hover_color="#2D3748",
        border_width=1,
        border_color="#4B5563",
        text_color="#9CA3AF",
        corner_radius=10,
        font=("Segoe UI", 12),
        height=40,
        command=cancelar
    )
    cancel_btn.pack(side="right")
    
    # ===== EFECTO DE FOCO INICIAL =====
    def highlight_option(frame):
        original_color = frame.cget("fg_color")
        frame.configure(fg_color="#2E313C")
        win.after(200, lambda: highlight_option(option2_frame))

def descargar_informes_inventario():
    print("🏁 Iniciando descarga de informes de inventario...")
    mostrar_toast("Iniciando descarga de informes...", tipo="info", titulo="Descarga")
    
    def _proceso_descarga():
        try:
            print("🔗 Abriendo sistema de inventario en Edge...")
            subprocess.Popen(["C:/Program Files (x86)/Microsoft/Edge/Application/msedge.exe", URL_RETIRADA])
            time.sleep(12)
            
            # Limpiar descargas anteriores
            limpiar_carpeta_descargas()
            
            # Crear carpeta FINAL desde el inicio
            carpeta_final = crear_carpeta_inventario()
            print(f"📁 TODOS los archivos se guardarán en: {carpeta_final}")
            
            carpeta_descargas = obtener_carpeta_descargas()
            archivos_descargados = []
            
            # Proceso para descargar 9 informes (uno por cada clínica)
            for i, codigo_clinica in enumerate(CODIGOS_CLINICAS):
                print(f"\n{'='*50}")
                print(f"📥 DESCARGANDO INFORME {i+1}/9 - CLÍNICA {codigo_clinica}")
                print(f"{'='*50}")
                
                # Buscar y hacer clic en el campo de unidad
                if not buscar_y_click("unidad_select.png", "unidad_select", confianza=0.6, intentos=3):
                    print(f"❌ No se pudo encontrar el campo 'unidad' para clínica {codigo_clinica}")
                    continue
                
                try:
                    time.sleep(2)
                    pyautogui.typewrite(codigo_clinica)
                    time.sleep(3)
                    x, y = pyautogui.position()
                    pyautogui.moveTo(x, y + 50)
                    pyautogui.click()
                    print(f"✅ Código de clínica {codigo_clinica} ingresado correctamente.")
                except Exception as e:
                    print(f"❌ Error al ingresar código de clínica {codigo_clinica}: {e}")
                    continue
                
                time.sleep(3)
                
                # Llenar los campos necesarios para el informe
                if not buscar_y_click("tipo_costo.png", "tipo_costo", intentos=3):
                    print(f"❌ No se encontró 'tipo_costo.png' para clínica {codigo_clinica}")
                    continue
                
                time.sleep(2)
                pyautogui.typewrite("01")
                time.sleep(3)
                x, y = pyautogui.position()
                pyautogui.moveTo(x, y + 40)
                pyautogui.click()
                time.sleep(1)
                
                if not buscar_y_click("nivel_totalizacion.png", "nivel_totalizacion", confianza=0.6, intentos=3):
                    print(f"❌ No se encontró 'nivel_totalizacion.png' para clínica {codigo_clinica}")
                    continue
                
                time.sleep(3)
                x, y = pyautogui.position()
                pyautogui.moveTo(x, y + 40)
                pyautogui.click()
                time.sleep(1)
                
                if not buscar_y_click("producto_inicial.png", "producto_inicial", intentos=3):
                    print(f"❌ No se encontró 'producto_inicial.png' para clínica {codigo_clinica}")
                    continue
                
                time.sleep(2)
                pyautogui.typewrite("1")
                time.sleep(3)
                x, y = pyautogui.position()
                pyautogui.moveTo(x, y + 40)
                pyautogui.click()
                time.sleep(1)
                
                if not buscar_y_click("producto_final.png", "producto_final", intentos=3):
                    print(f"❌ No se encontró 'producto_final.png' para clínica {codigo_clinica}")
                    continue
                
                time.sleep(2)
                pyautogui.typewrite("5")
                time.sleep(3)
                x, y = pyautogui.position()
                pyautogui.moveTo(x, y + 40)
                pyautogui.click()
                time.sleep(1)
                
                if fecha_personalizada:
                    fecha_sin_slash = fecha_personalizada.replace("/", "")
                    buscar_y_click("fecha_campo.png", "campo_fecha", confianza=0.7)
                    pyautogui.press("backspace", presses=10)
                    pyautogui.typewrite(fecha_sin_slash)
                    time.sleep(1)
                
                # Buscar y hacer clic en el botón de descarga XLSX
                if not buscar_y_click("generar_xlsx.png", "generar_xlsx", confianza=0.9, intentos=3):
                    print(f"❌ No se encontró 'generar_xlsx.png' para clínica {codigo_clinica}")
                    continue
                
                print(f"✅ Descarga iniciada para clínica {codigo_clinica}")
                
                # Esperar a que se complete la descarga
                print("⏳ Esperando a que se complete la descarga...")
                time.sleep(3)
                
                # Buscar el archivo descargado
                archivo_descargado = esperar_archivo_descargado(carpeta_descargas, tiempo_maximo=5)
                
                if archivo_descargado:
                    # Mover DIRECTAMENTE a la carpeta final
                    archivo_renombrado = renombrar_y_mover_archivo(
                        archivo_descargado, 
                        carpeta_final,  # ¡Directo a la carpeta final!
                        codigo_clinica, 
                        i
                    )
                    if archivo_renombrado:
                        archivos_descargados.append(archivo_renombrado)
                        print(f"✅ ÉXITO: Archivo guardado en carpeta final: {archivo_renombrado}")
                    else:
                        print(f"❌ FALLO: No se pudo mover archivo para clínica {codigo_clinica}")
                else:
                    print(f"❌ FALLO: No se descargó archivo para clínica {codigo_clinica}")
                
                # Recargar la página para la siguiente clínica (solo si no es la última)
                if i < len(CODIGOS_CLINICAS) - 1:
                    print("🔄 Recargando página para siguiente clínica...")
                    try:
                        pyautogui.hotkey("ctrl", "r")
                        time.sleep(7)
                    except Exception:
                        try:
                            pyautogui.press("f5")
                            time.sleep(7)
                        except Exception:
                            pass
            
            # RESUMEN FINAL
            print(f"\n{'='*60}")
            print("🎯 RESUMEN DE DESCARGA")
            print(f"{'='*60}")
            print(f"📁 Carpeta final: {carpeta_final}")
            print(f"📊 Archivos descargados: {len(archivos_descargados)}/9")
            
            if archivos_descargados:
                print("✅ Descarga completada exitosamente")
                mostrar_toast(f"Descarga completada: {len(archivos_descargados)}/9 archivos", tipo="success", titulo="Éxito")
                
                # Actualizar interfaz
                root.after(0, lambda: actualizar_info_inventario(
                    f"Carpeta: {os.path.basename(carpeta_final)}", 
                    f"Archivos: {len(archivos_descargados)}/9"
                ))
                
                # Habilitar botón de procesar
                root.after(0, lambda: btn_procesar_informes.configure(state="normal"))
                
                # Maximizar ventana
                root.after(0, lambda: root.state('zoomed'))
                root.after(0, lambda: root.deiconify())
            else:
                print("❌ No se descargó ningún archivo")
                mostrar_toast("No se descargaron archivos. Revisa el proceso.", tipo="warning", titulo="Advertencia")
                
        except Exception as e:
            print(f"❌ Error en descarga de informes: {e}")
            mostrar_toast(f"Error en descarga de informes:\n{e}", tipo="error", titulo="Error")
        finally:
            # Reactivar botón de descargar
            root.after(0, lambda: btn_descargar_informes.configure(state="normal"))
            print("🔄 Botón 'Descargar Informes' REACTIVADO")
    
    btn_descargar_informes.configure(state="disabled")
    threading.Thread(target=_proceso_descarga, daemon=True).start()

def procesar_informes_inventario():
    """
    Procesa los informes ya descargados en la carpeta final
    y genera un archivo resumen.
    """
    print("🔄 Procesando informes descargados...")
    mostrar_toast("Generando resumen...", tipo="info", titulo="Procesamiento")
    
    def _proceso():
        try:
            # Buscar la carpeta más reciente en el escritorio
            escritorio = os.path.join(os.path.expanduser("~"), "Escritorio")
            if not os.path.isdir(escritorio):
                escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
            
            # Buscar carpetas que comiencen con "Inventario"
            carpetas_inventario = []
            for item in os.listdir(escritorio):
                ruta = os.path.join(escritorio, item)
                if os.path.isdir(ruta) and item.startswith("Inventario"):
                    carpetas_inventario.append((ruta, os.path.getmtime(ruta)))
            
            if not carpetas_inventario:
                mostrar_toast("No se encontró carpeta de inventario", tipo="warning", titulo="Advertencia")
                return
            
            # Usar la carpeta más reciente
            carpetas_inventario.sort(key=lambda x: x[1], reverse=True)
            carpeta_final = carpetas_inventario[0][0]
            print(f"📁 Procesando carpeta: {carpeta_final}")
            
            # Encontrar archivos Excel en la carpeta
            archivos_excel = []
            for archivo in os.listdir(carpeta_final):
                if archivo.endswith(".xlsx") and not archivo.startswith("Resumen"):
                    archivos_excel.append(os.path.join(carpeta_final, archivo))
            
            if not archivos_excel:
                mostrar_toast("No hay informes para procesar", tipo="warning", titulo="Advertencia")
                return
            
            # Extraer la fecha del nombre de la carpeta
            nombre_carpeta = os.path.basename(carpeta_final)
            fecha_match = re.search(r'(\d{2}-\d{2}-\d{4})', nombre_carpeta)
            fecha_inventario = fecha_match.group(1) if fecha_match else datetime.now().strftime("%d-%m-%Y")
            
            # Procesar cada archivo
            resumen = []
            for archivo in archivos_excel:
                try:
                    df = pd.read_excel(archivo, header=None)
                    nombre_clinica = os.path.basename(archivo).replace(".xlsx", "")
                    
                    # Buscar total del stock (lógica existente)
                    total_stock = 0
                    patrones = ["total del stock", "valor total del stock", "total stock", "stock total"]
                    
                    for i in range(len(df)):
                        for j in range(len(df.columns)):
                            try:
                                celda = str(df.iat[i, j]).lower()
                            except:
                                celda = ""
                            if any(p in celda for p in patrones):
                                try:
                                    valor_bruto = df.iat[i, j + 2]
                                    valor_float = float(str(valor_bruto).replace(",", "."))
                                    total_stock = int(valor_float)
                                except:
                                    total_stock = 0
                                break
                    
                    resumen.append([fecha_inventario, nombre_clinica, total_stock, ""])
                    
                except Exception as e:
                    print(f"❌ Error procesando {archivo}: {e}")
                    continue
            
            # Crear DataFrame y guardar resumen
            df_final = pd.DataFrame(resumen, columns=["fecha", "clinica", "inventario", "compra"])
            
            # Usar 'ruta_resumen' no 'ruta_salida'
            nombre_resumen = f"Resumen Inventario {fecha_inventario}.xlsx"
            ruta_resumen = os.path.join(carpeta_final, nombre_resumen)  # ¡Aquí se define ruta_resumen!
            
            # Guardar con formato
            df_final.to_excel(ruta_resumen, index=False)
            
            # Aplicar formato al Excel
            wb = openpyxl.load_workbook(ruta_resumen)  # Usar ruta_resumen
            ws = wb.active
            
            # Formato del encabezado
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col in range(1, ws.max_column + 1):
                c = ws.cell(row=1, column=col)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center")
            
            # Bordes para todas las celdas
            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
            
            # Formato de moneda para la columna de inventario
            for row in range(2, ws.max_row + 1):
                ws[f"C{row}"].number_format = '"$"#,##0'
            
            # Autoajuste de columnas
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
            
            # Filtros automáticos
            ws.auto_filter.ref = f"A1:D{ws.max_row}"
            
            # Guardar cambios
            wb.save(ruta_resumen)  # Usar ruta_resumen
            
            print(f"📁 Resumen generado: {ruta_resumen}")
            mostrar_toast(f"Resumen generado en la carpeta {ruta_resumen}", tipo="success", titulo="Completado")
            
            # Abrir la carpeta
            try:
                os.startfile(carpeta_final)
            except:
                pass
            
        except Exception as e:
            print(f"❌ Error general: {e}")
            mostrar_toast(f"Error procesando inventario: {e}", tipo="error", titulo="Error")
        finally:
            root.after(0, lambda: btn_procesar_informes.configure(state="normal"))
    
    btn_procesar_informes.configure(state="disabled")
    threading.Thread(target=_proceso, daemon=True).start()

def actualizar_info_inventario(nombre_archivo="", total_items=0):
    try:
        lbl_info_inventario.configure(text=f"Archivo creado: {nombre_archivo}")
        lbl_estado_inventario.configure(text=f"Total de items procesados: {total_items}")
    except Exception as e:
        print(f"Error actualizando info inventario: {e}")

def actualizar_tabla_inventario():
    pass

# FUNCIONES AUXILIARES ORIGINALES 

def buscar_y_click(imagen, nombre, confianza=0.9, intentos=3, esperar=5):
    # Resolver ruta absoluta (CLAVE)
    ruta_imagen = os.path.abspath(os.path.join(CARPETA_BOTONES, imagen))
    inicio_total = time.time()

    # Validación temprana (evita 100 warnings inútiles)
    if not os.path.isfile(ruta_imagen):
        print(f"❌ Imagen no encontrada: {ruta_imagen}")
        return False

    pyautogui.FAILSAFE = True
    pyautogui.PAUSE = 0.1

    for i in range(intentos):
        print(f"🔎 Buscando botón '{nombre}' (Intento {i+1}/{intentos})...")

        tiempo_inicio_intento = time.time()
        ubicacion_estable = None
        frames_estables = 0

        while time.time() - tiempo_inicio_intento < esperar:
            try:
                ubicacion = pyautogui.locateCenterOnScreen(
                    ruta_imagen,
                    confidence=max(confianza - 0.15, 0.6),
                    grayscale=True
                )

                if ubicacion:
                    if ubicacion_estable:
                        dx = abs(ubicacion.x - ubicacion_estable.x)
                        dy = abs(ubicacion.y - ubicacion_estable.y)

                        if dx < 4 and dy < 4:
                            frames_estables += 1
                        else:
                            frames_estables = 1
                            ubicacion_estable = ubicacion
                    else:
                        ubicacion_estable = ubicacion
                        frames_estables = 1

                    if frames_estables >= 3:
                        pyautogui.moveTo(ubicacion_estable, duration=0.25)
                        pyautogui.click()
                        print(f"✅ Botón '{nombre}' encontrado y clickeado.")
                        return True
                else:
                    frames_estables = 0
                    ubicacion_estable = None

            except Exception:
                pass

            time.sleep(0.3)

        print("⚠️ No encontrado en este intento, reintentando...")

    print(
        f"❌ No se encontró el botón '{nombre}' tras {intentos} intentos "
        f"({int(time.time() - inicio_total)}s)."
    )
    return False

# ============================================================
# FUNCIONES DE AUTOMATIZACIÓN DESDE SUPABASE
# ============================================================

def obtener_facturas_pendientes_supabase():
    """Obtiene facturas con estado_procesamiento = 'pendiente'"""
    try:
        print("\n" + "="*60)
        print("🔍 DEBUG: Buscando facturas pendientes en Supabase...")
        print(f"   Estado buscado: '{ESTADO_PENDIENTE}'")
        print(f"   Max intentos: {MAX_INTENTOS}")
        print("="*60)
        
        # PRUEBA 1: Contar TODAS las facturas
        response_total = supabase.table('facturas').select('id', count='exact').execute()
        print(f"📊 Total de facturas en la tabla: {response_total.count}")
        
        # PRUEBA 2: Contar facturas con estado_procesamiento = 'pendiente'
        response_pendientes = supabase.table('facturas') \
            .select('id', count='exact') \
            .eq('estado_procesamiento', 'pendiente') \
            .execute()
        print(f"📊 Facturas con estado 'pendiente': {response_pendientes.count}")
        
        # PRUEBA 3: Ver los valores REALES de estado_procesamiento
        response_estados = supabase.table('facturas') \
            .select('id, numero_factura, estado_procesamiento, intentos_procesamiento') \
            .limit(5) \
            .execute()
        
        print(f"\n📋 Muestra de estados reales en BD:")
        for f in response_estados.data:
            print(f"   ID: {f['id']} | Factura: {f['numero_factura']} | Estado: '{f.get('estado_procesamiento')}' | Intentos: {f.get('intentos_procesamiento')}")
        
        # CONSULTA ORIGINAL (con debugging)
        print(f"\n🔎 Ejecutando consulta original...")
        response = supabase.table('facturas') \
            .select('id, numero_factura, pdf_url, codigo_unidad, valor_total, fecha_factura, intentos_procesamiento, proveedores(nombre, nit)') \
            .eq('estado_procesamiento', ESTADO_PENDIENTE) \
            .lt('intentos_procesamiento', MAX_INTENTOS) \
            .order('created_at', desc=False) \
            .execute()
        
        facturas = response.data
        print(f"✅ Resultado: {len(facturas)} facturas encontradas con la consulta completa")
        
        # Si no encontró nada, probar SIN el filtro de intentos
        if len(facturas) == 0:
            print(f"\n⚠️ No se encontraron facturas con la consulta completa")
            print(f"🔄 Probando SIN el filtro de intentos_procesamiento...")
            
            response_sin_filtro = supabase.table('facturas') \
                .select('id, numero_factura, pdf_url, codigo_unidad, valor_total, fecha_factura, intentos_procesamiento, proveedores(nombre, nit)') \
                .eq('estado_procesamiento', 'pendiente') \
                .order('created_at', desc=False) \
                .execute()
            
            facturas_sin_filtro = response_sin_filtro.data
            print(f"📊 Facturas encontradas SIN filtro de intentos: {len(facturas_sin_filtro)}")
            
            if len(facturas_sin_filtro) > 0:
                print(f"⚠️ PROBLEMA DETECTADO: El filtro de 'intentos_procesamiento' está bloqueando")
                print(f"   Usando facturas sin filtro de intentos...")
                facturas = facturas_sin_filtro
        
        print(f"\n{'='*60}")
        return facturas
    
    except Exception as e:
        print(f"❌ Error al obtener facturas: {e}")
        import traceback
        print(f"Stack trace completo:")
        traceback.print_exc()
        mostrar_toast(f"Error: {e}", tipo="error", titulo="Error")
        return []


def descargar_pdf_desde_supabase(pdf_url, factura_id):
    """Descarga un PDF desde Supabase Storage"""
    try:
        print(f"📥 Descargando PDF de factura {factura_id}...")
        print(f"   URL: {pdf_url[:80]}...")

        response = requests.get(pdf_url, timeout=30)
        response.raise_for_status()

        # Crear la ruta completa del archivo
        pdf_path = TEMP_PDF_DIR / f"factura_{factura_id}.pdf"

        # Escribir el contenido del PDF
        with open(pdf_path, 'wb') as f:
            f.write(response.content)

        # Verificar que el archivo fue creado
        if pdf_path.exists():
            tamaño = pdf_path.stat().st_size
            print(f"✅ PDF descargado exitosamente: {pdf_path}")
            print(f"   Tamaño: {tamaño:,} bytes")
        else:
            raise Exception("El archivo no se creó en el disco")

        return str(pdf_path.absolute())

    except Exception as e:
        print(f"❌ Error al descargar PDF: {e}")
        raise

import os
print("📂 Directorio actual de ejecución:", os.getcwd())


def limpiar_pdfs_temporales():
    """Limpia los PDFs descargados de la carpeta temporal"""
    try:
        archivos_eliminados = 0
        if TEMP_PDF_DIR.exists():
            for archivo in TEMP_PDF_DIR.glob("factura_*.pdf"):
                try:
                    archivo.unlink()
                    archivos_eliminados += 1
                    print(f"🗑️ Eliminado: {archivo.name}")
                except Exception as e:
                    print(f"⚠️ No se pudo eliminar {archivo.name}: {e}")

        print(f"🧹 Total de PDFs eliminados: {archivos_eliminados}")
        return archivos_eliminados
    except Exception as e:
        print(f"❌ Error limpiando PDFs: {e}")
        return 0


def actualizar_estado_factura_supabase(factura_id, estado, error=None):
    """Actualiza el estado de procesamiento en la BD"""
    try:
        factura = supabase.table('facturas') \
            .select('intentos_procesamiento') \
            .eq('id', factura_id) \
            .execute()
        
        intentos_actual = factura.data[0]['intentos_procesamiento'] if factura.data else 0
        
        datos = {
            'estado_procesamiento': estado,
            'intentos_procesamiento': intentos_actual + 1
        }
        
        if estado == ESTADO_COMPLETADO:
            from datetime import datetime
            datos['fecha_procesamiento'] = datetime.now().isoformat()
        
        if error:
            datos['error_procesamiento'] = error
        
        supabase.table('facturas').update(datos).eq('id', factura_id).execute()
        print(f"✅ Estado actualizado a '{estado}'")
    
    except Exception as e:
        print(f"❌ Error al actualizar estado: {e}")


def procesar_facturas_pendientes_automatico():
    """FUNCIÓN PRINCIPAL - Procesa facturas pendientes"""
    print("\n" + "="*60)
    print("🤖 PROCESAMIENTO AUTOMÁTICO DE FACTURAS")
    print("="*60)
    
    mostrar_toast("Buscando facturas pendientes...", tipo="info", titulo="Procesamiento")
    
    def _proceso():
        try:
            facturas = obtener_facturas_pendientes_supabase()
            
            if not facturas:
                root.after(0, lambda: mostrar_toast(
                    "No hay facturas pendientes",
                    tipo="info",
                    titulo="Sin facturas"
                ))
                return
            
            root.after(0, lambda: mostrar_toast(
                f"Procesando {len(facturas)} facturas...",
                tipo="info",
                titulo="Procesamiento"
            ))
            
            exitosas = 0
            fallidas = 0
            
            for i, factura in enumerate(facturas):
                print(f"\n📋 Factura {i+1}/{len(facturas)}")
                
                factura_id = factura['id']
                numero_factura = factura['numero_factura']
                pdf_url = factura.get('pdf_url')
                pdf_path = None
                
                try:
                    # Actualizar a "procesando"
                    actualizar_estado_factura_supabase(factura_id, ESTADO_PROCESANDO)
                    
                    # Descargar PDF
                    if pdf_url:
                        pdf_path = descargar_pdf_desde_supabase(pdf_url, factura_id)
                    else:
                        raise Exception("No hay PDF en Supabase")
                    
                    # AQUÍ INTEGRAR TU LÓGICA DE AUTOMATIZACIÓN EXISTENTE
                    # Por ahora solo simulamos
                    print(f"🤖 Procesando con PDF: {pdf_path}")
                    time.sleep(1)
                    
                    # Actualizar a "completado"
                    actualizar_estado_factura_supabase(factura_id, ESTADO_COMPLETADO)
                    exitosas += 1

                    # NO eliminar PDFs - mantener archivos descargados
                    # if pdf_path and os.path.exists(pdf_path):
                    #     os.remove(pdf_path)
                    print(f"📄 PDF conservado en: {pdf_path}")
                    
                except Exception as e:
                    print(f"❌ Error: {e}")
                    actualizar_estado_factura_supabase(factura_id, ESTADO_ERROR, str(e))
                    fallidas += 1
            
            # Resumen
            print(f"\n✅ Exitosas: {exitosas}\n❌ Fallidas: {fallidas}")
            root.after(0, lambda: mostrar_toast(
                f"Completado\n✅ {exitosas} exitosas\n❌ {fallidas} fallidas",
                tipo="success" if fallidas == 0 else "warning",
                titulo="Completado",
                duracion=5000
            ))
            
        except Exception as e:
            print(f"❌ Error general: {e}")
            root.after(0, lambda: mostrar_toast(
                f"Error: {e}",
                tipo="error",
                titulo="Error"
            ))
    
    threading.Thread(target=_proceso, daemon=True).start()



def cargar_datos_desde_supabase():
    global df_facturas, productos_por_factura, codigo_clinica, origen_datos

    try:
        print("\n" + "="*60)
        print("🔄 Cargando datos desde Supabase...")
        print("="*60)
        
        # Obtener facturas con proveedores
        facturas_response = supabase.table("facturas") \
            .select("*, proveedores(nombre, nit)") \
            .execute()
        
        # Obtener items con JOIN a catalogo_productos
        items_response = supabase.table("factura_items") \
            .select("*, catalogo_productos(codigo_arbol, nombre)") \
            .execute()

        print(f"📊 Facturas obtenidas: {len(facturas_response.data)}")
        print(f"📦 Items obtenidos: {len(items_response.data)}")

        if not facturas_response.data:
            mostrar_toast("No se encontraron facturas en Supabase", tipo="warning", titulo="Sin datos")
            return
        
        if not items_response.data:
            mostrar_toast("No se encontraron items en Supabase", tipo="warning", titulo="Sin datos")
            return

        facturas = []
        productos_por_factura = {}

        # Obtener código de clínica del primer registro
        if facturas_response.data:
            codigo_clinica = str(facturas_response.data[0].get("codigo_unidad", "0000")).strip()
            if not codigo_clinica or codigo_clinica == "None":
                codigo_clinica = "0000"
        
        print(f"🏥 Código de clínica detectado: {codigo_clinica}")

        # Procesar facturas
        for f in facturas_response.data:
            factura_id = str(f["id"])
            numero_factura = str(f["numero_factura"])
            
            # Obtener datos del proveedor
            proveedor = f.get("proveedores", {})
            if proveedor is None:
                proveedor = {}
            
            nombre_proveedor = proveedor.get("nombre", "Sin nombre") if isinstance(proveedor, dict) else "Sin nombre"
            nit_proveedor = proveedor.get("nit", "Sin NIT") if isinstance(proveedor, dict) else "Sin NIT"
            
            facturas.append({
                "ID_Factura": factura_id,
                "N° Factura": numero_factura,
                "Fecha": str(f["fecha_factura"]),
                "Empresa": nombre_proveedor,
                "NIT": nit_proveedor
            })
            
            # Inicializar lista de productos para esta factura
            productos_por_factura[numero_factura] = []
            
            print(f"   ✅ Factura {numero_factura}: {nombre_proveedor} (NIT: {nit_proveedor})")

        # Crear mapeo de ID a número de factura
        facturas_por_id = {
            str(f["id"]).strip(): str(f["numero_factura"]).strip()
            for f in facturas_response.data
        }

        # Procesar items (productos de cada factura)
        productos_agregados = 0
        for item in items_response.data:
            factura_id = str(item.get("factura_id", "")).strip()
            if not factura_id:
                continue

            numero_factura = facturas_por_id.get(factura_id)
            if not numero_factura:
                print(f"   ⚠️ Item sin factura válida: factura_id={factura_id}")
                continue

            if numero_factura not in productos_por_factura:
                productos_por_factura[numero_factura] = []



            # Obtener datos del catálogo de productos (JOIN)
            catalogo = item.get("catalogo_productos", {})
            if catalogo is None:
                catalogo = {}

            codigo_producto = catalogo.get("codigo_arbol", "SIN CODIGO") if isinstance(catalogo, dict) else "SIN CODIGO"
            nombre_producto = catalogo.get("nombre", "SIN NOMBRE") if isinstance(catalogo, dict) else "SIN NOMBRE"

            productos_por_factura[numero_factura].append({
                "Código Producto": str(codigo_producto).strip(),
                "Nombre Producto": str(nombre_producto).strip(),
                "Cantidad": str(item.get("cantidad", "")).replace(",", ".").strip(),
                "Precio": str(item.get("precio_unitario", "")).replace(",", ".").strip(),
            })


            productos_agregados += 1

        df_facturas = pd.DataFrame(facturas)
        origen_datos = "supabase"

        print(f"\n{'='*60}")
        print(f"✅ RESUMEN DE CARGA")
        print(f"{'='*60}")
        print(f"📋 Facturas cargadas: {len(facturas)}")
        print(f"📦 Productos cargados: {productos_agregados}")
        print(f"🏥 Código clínica: {codigo_clinica}")
        print(f"{'='*60}\n")

        # Mostrar resumen por factura
        for k, v in productos_por_factura.items():
            print(f"   📄 Factura {k}: {len(v)} productos")

        # Actualizar interfaz
        actualizar_tabla_facturas()
        
        try:
            btn_iniciar.configure(state="normal")
            quick_pulse_animation(btn_iniciar, "#4CAF50")
        except Exception:
            pass
        
        try:
            lbl_codigo.configure(text=f"🏥 Datos cargados desde Supabase (Clínica: {codigo_clinica})")
        except Exception:
            pass

        mostrar_toast(
            f"✅ {len(facturas)} facturas y {productos_agregados} productos cargados",
            tipo="success",
            titulo="Datos cargados"
        )

    except Exception as e:
        print(f"❌ Error al cargar datos desde Supabase: {e}")
        import traceback
        traceback.print_exc()
        mostrar_toast(f"Error al conectar con Supabase:\n{e}", tipo="error", titulo="Error")

def cargar_datos_desde_excel():
    archivo = filedialog.askopenfilename(
        title="Seleccionar archivo Excel",
        filetypes=[("Archivos Excel", "*.xlsx *.xls")],
    )
    if not archivo:
        mostrar_toast("No se seleccionó ningún archivo.", tipo="warning", titulo="Aviso")
        return

    global archivo_excel, df_facturas, productos_por_factura, codigo_clinica, origen_datos
    archivo_excel = archivo

    try:
        df = pd.read_excel(archivo, dtype=str).fillna("")
        columnas_esperadas = [
            "Tipo",
            "N° Factura",
            "Fecha",
            "Empresa",
            "NIT",
            "Código Producto",
            "Nombre Producto",
            "Cantidad",
            "Precio",
        ]
        for col in columnas_esperadas:
            if col not in df.columns:
                mostrar_toast(f"Falta la columna '{col}' en el archivo.", tipo="error", titulo="Error de formato")
                return

        facturas = []
        productos_por_factura = {}
        factura_actual = None

        for _, fila in df.iterrows():
            tipo = str(fila["Tipo"]).strip().upper()

            if tipo == "FACTURA":
                factura_actual = str(fila["N° Factura"]).strip()
                if not factura_actual:
                    continue
                facturas.append(
                    {
                        "ID_Factura": factura_actual,
                        "N° Factura": factura_actual,
                        "Fecha": fila["Fecha"].strip(),
                        "Empresa": fila["Empresa"].strip(),
                        "NIT": fila["NIT"].strip(),
                    }
                )
                productos_por_factura[factura_actual] = []
            elif tipo == "PRODUCTO" and factura_actual:
                productos_por_factura[factura_actual].append(
                    {
                        "Código Producto": str(fila["Código Producto"]).strip(),
                        "Nombre Producto": str(fila["Nombre Producto"]).strip(),
                        "Cantidad": str(fila["Cantidad"]).replace(",", ".").strip(),
                        "Precio": str(fila["Precio"]).replace(",", ".").strip(),
                    }
                )

        df_facturas = pd.DataFrame(facturas)
        codigo_clinica = "".join(
            [c for c in os.path.basename(archivo) if c.isdigit()][-4:]
        )
        if not codigo_clinica:
            codigo_clinica = "0000"

        origen_datos = "excel"
        actualizar_tabla_facturas()
        try:
            btn_iniciar.configure(state="normal")
            quick_pulse_animation(btn_iniciar, "#4CAF50")
        except Exception:
            pass
        try:
            lbl_codigo.configure(text=f"🏥 Código clínica detectado: {codigo_clinica}")
        except Exception:
            pass
        mostrar_toast("Archivo cargado correctamente ✅", tipo="success", titulo="Éxito")

    except Exception as e:
        mostrar_toast(f"No se pudo leer el archivo:\n{e}", tipo="error", titulo="Error")

def actualizar_tabla_facturas():
    for row in tree_facturas.get_children():
        tree_facturas.delete(row)
    for _, fila in df_facturas.iterrows():
        tree_facturas.insert("", "end", values=list(fila))

def mostrar_productos(event):
    item = tree_facturas.selection()
    if not item:
        return
    valores = tree_facturas.item(item[0], "values")
    if len(valores) < 2:
        return
    numero_factura = valores[1]

    productos = productos_por_factura.get(numero_factura, [])
    for row in tree_productos.get_children():
        tree_productos.delete(row)

    for prod in productos:
        tree_productos.insert(
            "",
            "end",
            values=(
                prod["Código Producto"],
                prod["Nombre Producto"],
                prod["Cantidad"],
                prod["Precio"],
            ),
        )

    try:
        lbl_productos.configure(
            text=f"🛒 Productos de la factura seleccionada (Total: {len(productos)})"
        )
    except Exception:
        pass
    print(f"🧾 Factura seleccionada: {numero_factura}")
    print(f"📦 Productos encontrados: {len(productos)}")

def iniciar_proceso():
    if df_facturas.empty:
        mostrar_toast("Primero carga datos desde Excel o Supabase.", tipo="error", titulo="Error")
        return

    print(f"🏁 Iniciando proceso para la clínica {codigo_clinica} ({origen_datos.upper()})...")
    print("🔹 Abriendo TecFood en Edge...")
    subprocess.Popen(["C:/Program Files (x86)/Microsoft/Edge/Application/msedge.exe", URL_TECFOOD])
    time.sleep(20)

    # Bucle principal: una iteración por cada fila (factura) en df_facturas
    for index, factura_actual in df_facturas.iterrows():
        numero_factura = str(factura_actual["N° Factura"]).strip()
        empresa = str(factura_actual.get("Empresa", "")).strip()
        nit = str(factura_actual.get("NIT", "")).strip()
        productos = productos_por_factura.get(numero_factura, [])

        print(f"\n\n🔁 Procesando factura {index+1}/{len(df_facturas)} → {numero_factura}")
        print(f"   Empresa: {empresa}  NIT: {nit}  Productos: {len(productos)}")

        if not productos:
            print(f"⚠️ Sin productos para {numero_factura}, se salta.")
            continue
        

        if not buscar_y_click("unidad_select.png", "unidad_select", confianza=0.6):
            mostrar_toast("No se encontró el botón unidad_select.png", tipo="warning", titulo="Botón no encontrado")
            continue

        # Selección de clínica
        try:
            time.sleep(3)
            pyautogui.typewrite(codigo_clinica)
            time.sleep(2)
            x, y = pyautogui.position()
            pyautogui.moveTo(x, y + 50, duration=0.5)
            pyautogui.click()
            print(f"✅ Clínica {codigo_clinica} seleccionada correctamente.")
        except Exception as e:
            print("Error al seleccionar clínica:", e)
            continue
        pyautogui.press("tab")
        pyautogui.press("tab")
        pyautogui.press("tab")
        time.sleep(5)
        try:
            fecha = pd.to_datetime(factura_actual["Fecha"])
            fecha_formateada = fecha.strftime("%d%m%Y")
            pyautogui.typewrite(fecha_formateada, interval=0.1)
            pyautogui.typewrite(fecha_formateada, interval=0.1)
            pyautogui.press("tab")
            print(f"📅 Fecha ingresada: {fecha_formateada}")
        except Exception as e:
            print("⚠️ Error al procesar la fecha:", e)
            continue

        # Aplicar filtro
        time.sleep(5)
        if not (buscar_y_click("aplicar_filtro.png", "aplicar_filtro") or buscar_y_click("aplicar_filtro_en.png", "aplicar_filtro_en.png")):
            mostrar_toast("No se encontró el botón 'aplicar_filtro'.", tipo="warning", titulo="Botón no encontrado")
            continue

        # Añadir factura
        time.sleep(6)
        if not buscar_y_click("anadir.png", "anadir", confianza=1, intentos=3):
            mostrar_toast("No se encontró el botón 'anadir.png'.", tipo="warning", titulo="Botón no encontrado")
            continue
        time.sleep(6)
        buscar_y_click("seleccionar_archivo.png", "seleccionar")
        time.sleep(2)

        # Buscar y seleccionar PDF correspondiente a esta factura
        try:
            empresa_limpia = "".join(c for c in empresa if c.isalnum() or c.isspace()).lower()
            nit_limpio = "".join(c for c in nit if c.isalnum()).lower()
            carpeta_pdf = r"C:\Users\Maicol Hernandez\Documents\GitHub\PyHealthy\PDF"
            pdf_encontrado: Optional[str] = None
            for archivo in os.listdir(carpeta_pdf):
                nombre_archivo = archivo.lower()
                if (nit_limpio in nombre_archivo and any(palabra in nombre_archivo for palabra in empresa_limpia.split()) and archivo.endswith(".pdf")):
                    pdf_encontrado = os.path.join(carpeta_pdf, archivo)
                    break

            if pdf_encontrado:
                print(f"📄 PDF encontrado: {pdf_encontrado}")
                pyautogui.typewrite(pdf_encontrado)
                time.sleep(1)
                pyautogui.press("enter")
                print("✅ PDF seleccionado correctamente.")
            else:
                mostrar_toast(f"No se encontró un PDF para '{empresa}' y '{nit}'. Verifica carpeta.", tipo="warning", titulo="PDF no encontrado")
                print("⚠️ PDF no encontrado. Se continúa con la factura (si corresponde).")
                # si quieres saltar la factura si no hay PDF, usa: continue
        except Exception as e:
            mostrar_toast(f"No se pudo procesar la factura o buscar el PDF:\n{e}", tipo="error", titulo="Error")
            continue

        # Remitente y número de factura
        time.sleep(5)
        if not buscar_y_click("remitente.png", "remitente"):
            mostrar_toast("No se encontró 'remitente.png'.", tipo="warning", titulo="Botón no encontrado")
            continue
        pyautogui.typewrite(nit)
        pyautogui.sleep(3)
        pyautogui.press("tab")
        pyautogui.sleep(7)
        pyautogui.typewrite(numero_factura)
        pyautogui.press("tab")
        pyautogui.sleep(2)

        if not buscar_y_click("serie.png", "serie"):
            mostrar_toast("No se encontró 'serie.png'.", tipo="warning", titulo="Botón no encontrado")
            continue
        x, y = pyautogui.position()
        pyautogui.moveTo(x, y + 40, duration=0.5)
        pyautogui.click()
        if not buscar_y_click("emision.png", "emision"):
            mostrar_toast("No se encontró 'emision.png'.", tipo="warning", titulo="Botón no encontrado")
            continue
        x, y = pyautogui.position()
        pyautogui.moveTo(x + 50, y + 30, duration=0.5)
        pyautogui.click()
        pyautogui.press("backspace", presses=10)
        pyautogui.typewrite(fecha_formateada)

        if not buscar_y_click("valor.png", "valor", confianza=1, intentos=3):
            mostrar_toast("No se encontró 'valor.png'.", tipo="warning", titulo="Botón no encontrado")
            continue

        pyautogui.typewrite("0")
        pyautogui.press("tab")

        if not buscar_y_click("grabar.png", "grabar"):
            mostrar_toast("No se encontró 'grabar.png'.", tipo="warning", titulo="Botón no encontrado")
            continue

        pyautogui_sleep = getattr(pyautogui, "sleep", None)
        if pyautogui_sleep:
            pyautogui_sleep(12)
        else:
            time.sleep(12)

        if not buscar_y_click("productos.png", "productos"):
            mostrar_toast("No se encontró 'productos.png'.", tipo="warning", titulo="Botón no encontrado")
            continue

        time.sleep(6)

        # Agregar productos
        for i, producto in enumerate(productos):
            try:
                codigo_producto = str(producto["Código Producto"]).strip()
                cantidad = str(producto["Cantidad"]).strip().replace(",", ".")
                valor_unitario = str(producto["Precio"]).strip().replace(",", ".")

                print(f"➕ Agregando producto {i+1}/{len(productos)}: {codigo_producto}")
                if not buscar_y_click("anadir.png", "añadir", confianza=1, intentos=3):
                    mostrar_toast("No se encontró 'anadir.png'.", tipo="warning", titulo="Botón no encontrado")
                    raise Exception("boton_anadir_no_encontrado")
                time.sleep(11)

                if not buscar_y_click("anadir_producto.png", "añadir_producto"):
                    mostrar_toast("No se encontró 'anadir_producto.png'.", tipo="warning", titulo="Botón no encontrado")
                    raise Exception("boton_anadir_producto_no_encontrado")
                time.sleep(11)

                pyautogui.typewrite(codigo_producto)
                time.sleep(3)
                x, y = pyautogui.position()
                pyautogui.moveTo(x, y + 40, duration=0.5)
                pyautogui.click()
                time.sleep(3)

                if not buscar_y_click("cantidad.png", "cantidad"):
                    mostrar_toast("No se encontró 'cantidad.png'.", tipo="warning", titulo="Botón no encontrado")
                    raise Exception("boton_cantidad_no_encontrado")
                pyautogui.typewrite(cantidad)
                pyautogui.press("tab")
                time.sleep(5)

                pyautogui.typewrite(valor_unitario)
                pyautogui.press("tab")
                time.sleep(5)

                if not buscar_y_click("grabar.png", "grabar"):
                    mostrar_toast("No se encontró 'grabar.png'.", tipo="warning", titulo="Botón no encontrado")
                    raise Exception("boton_grabar_no_encontrado")

                print(f"✅ Producto {codigo_producto} grabado correctamente.")
                time.sleep(5)
                pyautogui.press("esc")
                pyautogui_sleep = getattr(pyautogui, "sleep", None)
                if pyautogui_sleep:
                    pyautogui_sleep(5)
                else:
                    time.sleep(5)
            except Exception as e:
                print(f"❌ Error al agregar producto {codigo_producto}: {e}")
                continue

        # Finalizar esta factura
        if not buscar_y_click("FinalizarF.png", "FinalizarF"):
            mostrar_toast("No se encontró 'FinalizarF.png'.", tipo="warning", titulo="Botón no encontrado")
            continue
        time.sleep(5)
        if not buscar_y_click("si.png", "si"):
            mostrar_toast("No se encontró 'si.png'.", tipo="warning", titulo="Botón no encontrado")
            continue
        time.sleep(2)
        
        print(f"✅ Factura {numero_factura} procesada correctamente (productos: {len(productos)})")

        # --- Recargar la página para la siguiente factura ---
        time.sleep(2)
        try:
            pyautogui.hotkey("ctrl", "r")
            time.sleep(10)
        except Exception:
            try:
                pyautogui.press("f5")
                time.sleep(10)
            except Exception:
                pass
    print("🏁 Proceso completado para todas las facturas.")

# INTERFAZ CON CUSTOMTKINTER
try:
    root
except NameError:
    silenciar_customtkinter()
    root = ctk.CTk()
    root.title("DataSpectra - Sistema de Inventario")
    root.attributes("-fullscreen", True)
    root.configure(fg_color="#0E0F12")

# Sistema simple de multipantallas
contenedor = ctk.CTkFrame(root, fg_color=BG)
contenedor.pack(fill="both", expand=True)

pantallas = {}
def mostrar_pantalla(nombre):
    for p in pantallas.values():
        p.pack_forget()
    pantallas[nombre].pack(fill="both", expand=True)

# PANTALLA DE LOGIN (nueva)
pantalla_login = ctk.CTkFrame(contenedor, fg_color=BG)
pantallas["login"] = pantalla_login

def cargar_interfaz_login():
    """Carga la interfaz de login dentro de la pantalla principal"""
    # Limpiar pantalla login si ya tiene contenido
    for widget in pantalla_login.winfo_children():
        widget.destroy()
    
    login_frame = ctk.CTkFrame(pantalla_login, fg_color=CARD_BG, corner_radius=20)
    login_frame.place(relx=0.5, rely=0.5, anchor="center")
    
    # Título
    title_label = ctk.CTkLabel(
        login_frame,
        text="🔐 DataSpectra",
        font=("Segoe UI", 26, "bold"),
        text_color=TEXT_MAIN
    )
    title_label.pack(pady=(40, 10))
    
    subtitle_label = ctk.CTkLabel(
        login_frame,
        text="Inicio de Sesión Obligatorio",
        font=("Segoe UI", 14),
        text_color=TEXT_SECOND
    )
    subtitle_label.pack(pady=(0, 30))
    
    # Formulario
    email_label = ctk.CTkLabel(
        login_frame,
        text="Correo Electrónico",
        font=("Segoe UI", 12),
        text_color=TEXT_MAIN,
        anchor="w"
    )
    email_label.pack(fill="x", padx=40, pady=(10, 5))
    
    email_entry = ctk.CTkEntry(
        login_frame,
        placeholder_text="usuario@ejemplo.com",
        width=300,
        height=45,
        font=("Segoe UI", 13),
        fg_color="#252833",
        border_color=PRIMARY,
        corner_radius=12
    )
    email_entry.pack(pady=(0, 15))
    
    password_label = ctk.CTkLabel(
        login_frame,
        text="Contraseña",
        font=("Segoe UI", 12),
        text_color=TEXT_MAIN,
        anchor="w"
    )
    password_label.pack(fill="x", padx=40, pady=(5, 5))
    
    password_entry = ctk.CTkEntry(
        login_frame,
        placeholder_text="••••••••",
        width=300,
        height=45,
        font=("Segoe UI", 13),
        fg_color="#252833",
        border_color=PRIMARY,
        corner_radius=12,
        show="•"
    )
    password_entry.pack(pady=(0, 10))
    
    status_label = ctk.CTkLabel(
        login_frame,
        text="",
        font=("Segoe UI", 12),
        text_color=TEXT_SECOND
    )
    status_label.pack(pady=(5, 10))
    
    def realizar_login_local():
        email = email_entry.get().strip()
        password = password_entry.get()
        
        if not email or not password:
            status_label.configure(text="Por favor completa todos los campos", text_color="#FF5252")
            return
        
        status_label.configure(text="Verificando credenciales...", text_color=ACCENT)
        
        # Deshabilitar temporalmente
        email_entry.configure(state="disabled")
        password_entry.configure(state="disabled")
        
        def login_thread():
            try:
                global supabase
                # Limpiar cualquier sesión previa
                try:
                    supabase.auth.sign_out()
                except:
                    pass
                
                # Intentar login
                auth_response = supabase.auth.sign_in_with_password({
                    "email": email,
                    "password": password
                })
                
                session = auth_response.session
                user = auth_response.user
                
                # Guardar sesión
                try:
                    with open("session.json", "w") as f:
                        json.dump({
                            "access_token": session.access_token,
                            "refresh_token": session.refresh_token,
                            "user_id": user.id,
                            "email": user.email
                        }, f)
                except Exception as e:
                    print(f"Error guardando sesión: {e}")
                
                # Mostrar éxito y cambiar a menú
                root.after(0, lambda: status_label.configure(
                    text="✓ Sesión iniciada correctamente", 
                    text_color="#4CAF50"
                ))
                
                root.after(1500, lambda: mostrar_pantalla("menu"))
                
            except Exception as e:
                error_msg = str(e)
                if "Invalid login credentials" in error_msg:
                    error_msg = "Credenciales inválidas"
                elif "Email not confirmed" in error_msg:
                    error_msg = "Email no confirmado"
                
                root.after(0, lambda: status_label.configure(
                    text=f"✗ Error: {error_msg}", 
                    text_color="#FF5252"
                ))
                
                # Rehabilitar campos
                root.after(0, lambda: email_entry.configure(state="normal"))
                root.after(0, lambda: password_entry.configure(state="normal"))
        
        threading.Thread(target=login_thread, daemon=True).start()
    
    # Botón de login
    login_button = ctk.CTkButton(
        login_frame,
        text="Iniciar Sesión",
        command=realizar_login_local,
        width=300,
        height=48,
        font=("Segoe UI", 15, "bold"),
        fg_color=PRIMARY,
        hover_color=PRIMARY_HOVER,
        corner_radius=12
    )
    login_button.pack(pady=(10, 20))
    
    # Footer
    footer_label = ctk.CTkLabel(
        login_frame,
        text="Solo usuarios registrados en Supabase",
        font=("Segoe UI", 11),
        text_color="#666"
    )
    footer_label.pack(pady=(10, 30))
    
    # Intentar auto-login
    def attempt_auto_login():
        try:
            if os.path.exists("session.json"):
                with open("session.json", "r") as f:
                    session_data = json.load(f)
                    if "access_token" in session_data:
                        status_label.configure(text="Recuperando sesión...", text_color=ACCENT)
                        
                        def verify_session():
                            try:
                                global supabase
                                user = supabase.auth.get_user(session_data["access_token"])
                                if user:
                                    root.after(0, lambda: status_label.configure(
                                        text="✓ Sesión recuperada", 
                                        text_color="#4CAF50"
                                    ))
                                    root.after(1000, lambda: mostrar_pantalla("menu"))
                                    return
                            except:
                                pass
                            # Si falla, limpiar sesión
                            try:
                                os.remove("session.json")
                            except:
                                pass
                        
                        threading.Thread(target=verify_session, daemon=True).start()
        except:
            pass
    
    root.after(500, attempt_auto_login)

# Crear la interfaz de login
cargar_interfaz_login()

# PANTALLA PRINCIPAL
pantalla_menu = ctk.CTkFrame(contenedor, fg_color=BG)
pantallas["menu"] = pantalla_menu

# Header menu
header_menu = ctk.CTkFrame(pantalla_menu, fg_color=CARD_BG, corner_radius=18)
header_menu.pack(fill="x", padx=20, pady=16)

titulo_menu = ctk.CTkLabel(header_menu, text="DataSpectra — Panel Principal", font=("Segoe UI", 22, "bold"), text_color=TEXT_MAIN)
titulo_menu.pack(side="left", padx=18, pady=10)

# Botón de logout
def cerrar_sesion():
    try:
        global supabase
        supabase.auth.sign_out()
        # Eliminar archivo de sesión
        if os.path.exists("session.json"):
            os.remove("session.json")
        mostrar_toast("Sesión cerrada correctamente", tipo="info", titulo="Cerrar sesión")
        mostrar_pantalla("login")
    except Exception as e:
        print(f"Error cerrando sesión: {e}")

btn_logout = ctk.CTkButton(header_menu, text="🚪 Salir", width=44, height=38, corner_radius=14,
                          fg_color="#FF5252", hover_color="#FF6B6B", font=("Segoe UI", 12, "bold"),
                          command=cerrar_sesion)
btn_logout.pack(side="right", padx=14, pady=8)

# Añadir hover simple
btn_logout.bind("<Enter>", lambda e: simple_button_hover(btn_logout, True))
btn_logout.bind("<Leave>", lambda e: simple_button_hover(btn_logout, False))

# Botón cerrar app
def cerrar_app_wrapper():
    if confirmar_salida("Salir", "¿Deseas cerrar la aplicación?"):
        mostrar_toast("Cerrando aplicación...", tipo="info", titulo="Cerrando")
        root.after(300, root.destroy)
    else:
        mostrar_toast("Salida cancelada", tipo="info", titulo="Cancelado")

btn_exit_menu = ctk.CTkButton(header_menu, text="✕", width=44, height=38, corner_radius=14,
                              fg_color="#FF5252", hover_color="#FF6B6B", font=("Segoe UI", 14, "bold"),
                              command=cerrar_app_wrapper)
btn_exit_menu.pack(side="right", padx=14, pady=8)

# Añadir hover simple
btn_exit_menu.bind("<Enter>", lambda e: simple_button_hover(btn_exit_menu, True))
btn_exit_menu.bind("<Leave>", lambda e: simple_button_hover(btn_exit_menu, False))

content_menu = ctk.CTkScrollableFrame(pantalla_menu, fg_color=BG, corner_radius=0)
content_menu.pack(fill="both", expand=True, padx=20, pady=(10,20))

intro = ctk.CTkLabel(content_menu, text="Selecciona un proceso", font=("Segoe UI", 28, "bold"), text_color=TEXT_MAIN)
intro.pack(pady=(12,6))
sub = ctk.CTkLabel(content_menu, text="Funciones disponibles (las futuras aparecerán habilitadas cuando el administrador las implemente).", font=("Segoe UI", 15), text_color=TEXT_SECOND)
sub.pack(pady=(0,12))

# Lista vertical
def crear_item_lista(parent, icon, titulo, subtitulo, command=None, enabled=True):
    item = ctk.CTkFrame(parent, fg_color="#141518", corner_radius=16)
    item.pack(fill="x", padx=12, pady=8)
    item.configure(height=92)
    item.grid_propagate(False)

    # Icono a la izquierda
    lbl_icon = ctk.CTkLabel(item, text=icon, font=("Segoe UI Emoji", 32), text_color=ACCENT, fg_color="#141518")
    lbl_icon.place(x=18, y=28)

    # Textos
    lbl_tit = ctk.CTkLabel(item, text=titulo, font=("Segoe UI", 15, "bold"), text_color=TEXT_MAIN, fg_color="#141518")
    lbl_tit.place(x=80, y=18)
    lbl_sub = ctk.CTkLabel(item, text=subtitulo, font=("Segoe UI", 11), text_color=TEXT_SECOND, fg_color="#141518")
    lbl_sub.place(x=80, y=44)

    #indicador
    lbl_chev = ctk.CTkLabel(item, text="›", font=("Segoe UI", 22, "bold"), text_color="#8F9AA6", fg_color="#141518")
    lbl_chev.place(relx=0.96, rely=0.5, anchor="e")

    # Hover
    def on_enter(e):
        if enabled:
            simple_item_hover(item, True)
    def on_leave(e):
        if enabled:
            simple_item_hover(item, False)
    item.bind("<Enter>", on_enter)
    item.bind("<Leave>", on_leave)

    # Click
    if enabled and callable(command):
        def onclick(e=None):
            try:
                quick_pulse_animation(item, ACCENT)
                mostrar_toast(f"Abriendo: {titulo}", tipo="info", titulo="Abriendo Operación")
                command()
            except Exception as err:
                mostrar_toast(f"Ocurrió un error: {err}", tipo="error", titulo="Error")
        item.bind("<Button-1>", onclick)
        lbl_icon.bind("<Button-1>", onclick)
        lbl_tit.bind("<Button-1>", onclick)
        lbl_sub.bind("<Button-1>", onclick)
        lbl_chev.bind("<Button-1>", onclick)
    else:
        lbl_tit.configure(text_color="#6F777E")
        lbl_sub.configure(text_color="#4F5559")
        lbl_chev.configure(text_color="#3B4043")

    return item

# Crear 3 items
crear_item_lista(content_menu, "🧾", "Cargar facturas", "Automatiza la carga de facturas en TecFood", command=lambda: mostrar_pantalla("facturas"), enabled=True)
crear_item_lista(content_menu, "📦", "Sistema de Inventario Automatizado", "Descarga y consolida informes de inventario de stock", command=lambda: mostrar_pantalla("inventario"), enabled=True)
crear_item_lista(content_menu, "⚙️", "Función futura #2", "Implementación posterior", command=None, enabled=False)

# footer pequeño
footer_menu = ctk.CTkLabel(content_menu, text="Desarrollado por Area de desarrollo Healthy", font=("Segoe UI", 11), text_color=TEXT_SECOND)
footer_menu.pack(pady=18)

# =========================
# PANTALLA DE FACTURAS
# =========================
pantalla_facturas = ctk.CTkFrame(contenedor, fg_color=BG)
pantallas["facturas"] = pantalla_facturas

# === HEADER SUPERIOR ===
header = ctk.CTkFrame(pantalla_facturas, fg_color=CARD_BG, corner_radius=20)
header.pack(fill="x", padx=20, pady=12)

lbl_titulo = ctk.CTkLabel(
    header,
    text="DataSpectra - Carga de Facturas TecFood",
    font=("Segoe UI", 24, "bold"),
    text_color=TEXT_MAIN,
)
lbl_titulo.pack(side="left", padx=20, pady=12)

# Botón volver
def _on_volver():
    mostrar_toast("Volviendo al menú", tipo="info", titulo="Volver")
    mostrar_pantalla("menu")

btn_volver = ctk.CTkButton(
    header,
    text="↩ Volver",
    command=_on_volver,
    width=90,
    height=36,
    corner_radius=16,
    fg_color="#5EA0FF",
    hover_color="#7AB8FF",
    font=("Segoe UI", 13, "bold"),
)
btn_volver.pack(side="right", padx=10, pady=10)

# Añadir hover simple
btn_volver.bind("<Enter>", lambda e: simple_button_hover(btn_volver, True))
btn_volver.bind("<Leave>", lambda e: simple_button_hover(btn_volver, False))

# Botón cerrar
def cerrar_app_facturas():
    if confirmar_salida("Salir", "¿Deseas cerrar la aplicación?"):
        mostrar_toast("Cerrando aplicación...", tipo="info", titulo="Cerrando")
        root.after(300, root.destroy)
    else:
        mostrar_toast("Salida cancelada", tipo="info", titulo="Cancelado")

btn_close = ctk.CTkButton(
    header,
    text="✕",
    command=cerrar_app_facturas,
    width=40,
    height=40,
    corner_radius=20,
    fg_color="#FF5252",
    hover_color="#FF6B6B",
    font=("Segoe UI", 14, "bold"),
)
btn_close.pack(side="right", padx=10, pady=10)

# Añadir hover simple
btn_close.bind("<Enter>", lambda e: simple_button_hover(btn_close, True))
btn_close.bind("<Leave>", lambda e: simple_button_hover(btn_close, False))

# === CUERPO PRINCIPAL ===
main_frame = ctk.CTkScrollableFrame(pantalla_facturas, fg_color=BG, corner_radius=0)
main_frame.pack(fill="both", expand=True, padx=30, pady=10)

# --- BOTONES DE CARGA ---
btn_frame = ctk.CTkFrame(main_frame, fg_color=CARD_BG, corner_radius=25)
btn_frame.pack(pady=18)

def _on_cargar_excel():
    mostrar_toast("Selecciona un archivo para cargar", tipo="info", titulo="Cargar Excel")
    quick_pulse_animation(btn_excel, "#FFD54F")
    try:
        cargar_datos_desde_excel()
    except Exception as e:
        mostrar_toast(f"Error al ejecutar cargar_datos_desde_excel: {e}", tipo="error", titulo="Error")

def _on_cargar_supabase():
    mostrar_toast("Cargando desde Supabase...", tipo="info", titulo="Supabase")
    quick_pulse_animation(btn_supabase, "#4FC3F7")
    try:
        cargar_datos_desde_supabase()
    except Exception as e:
        mostrar_toast(f"Error al ejecutar cargar_datos_desde_supabase: {e}", tipo="error", titulo="Error")

btn_excel = ctk.CTkButton(
    btn_frame,
    text="📂 Cargar Excel",
    command=_on_cargar_excel,
    corner_radius=25,
    fg_color=ACCENT,
    hover_color=ACCENT_HOVER,
    text_color="#000000",
    font=("Segoe UI", 14, "bold"),
    width=200,
)
btn_excel.pack(side="left", padx=12, pady=12)

btn_supabase = ctk.CTkButton(
    btn_frame,
    text="🗄️ Cargar desde Supabase",
    command=_on_cargar_supabase,
    corner_radius=25,
    fg_color=PRIMARY,
    hover_color=PRIMARY_HOVER,
    text_color="#FFFFFF",
    font=("Segoe UI", 14, "bold"),
    width=250,
)
btn_supabase.pack(side="left", padx=12, pady=12)

btn_excel.bind("<Enter>", lambda e: simple_button_hover(btn_excel, True))
btn_excel.bind("<Leave>", lambda e: simple_button_hover(btn_excel, False))
btn_supabase.bind("<Enter>", lambda e: simple_button_hover(btn_supabase, True))
btn_supabase.bind("<Leave>", lambda e: simple_button_hover(btn_supabase, False))
# Botón de procesamiento automático
btn_automatico = ctk.CTkButton(
    btn_frame,
    text="🤖 Procesar Pendientes",
    command=procesar_facturas_pendientes_automatico,
    corner_radius=25,
    fg_color="#9C27B0",
    hover_color="#BA68C8",
    text_color="#FFFFFF",
    font=("Segoe UI", 14, "bold"),
    width=250,
)
btn_automatico.pack(side="left", padx=12, pady=12)

# Agregar hover
btn_automatico.bind("<Enter>", lambda e: btn_automatico.configure(fg_color="#BA68C8"))
btn_automatico.bind("<Leave>", lambda e: btn_automatico.configure(fg_color="#9C27B0"))

btn_supabase.pack(side="left", padx=12, pady=12)
# --- CÓDIGO DE CLÍNICA ---
lbl_codigo = ctk.CTkLabel(
    main_frame,
    text="Código de clínica detectado: ----",
    font=("Segoe UI", 14, "bold"),
    text_color=TEXT_SECOND,
)
lbl_codigo.pack(pady=6)

# === SECCIÓN DE TABLAS ===
tables_frame = ctk.CTkFrame(main_frame, fg_color=CARD_BG, corner_radius=20)
tables_frame.pack(fill="both", expand=True, padx=10, pady=10)

def crear_titulo(text):
    return ctk.CTkLabel(
        tables_frame,
        text=text,
        font=("Segoe UI", 16, "bold"),
        text_color=ACCENT,
        anchor="w",
    )

# --- FACTURAS ---
lbl_facturas = crear_titulo("📑 Facturas detectadas")
lbl_facturas.pack(anchor="w", pady=(10, 0), padx=12)

tree_facturas = ttk.Treeview(
    tables_frame,
    columns=("ID_Factura", "N° Factura", "Fecha", "Empresa", "NIT"),
    show="headings",
    height=8,
)
for col in ("ID_Factura", "N° Factura", "Fecha", "Empresa", "NIT"):
    tree_facturas.heading(col, text=col)
    tree_facturas.column(col, anchor="center", width=170)
tree_facturas.pack(fill="x", padx=12, pady=8)
if callable(globals().get("mostrar_productos")):
    tree_facturas.bind("<Double-1>", mostrar_productos)

# --- PRODUCTOS ---
lbl_productos = crear_titulo("🛒 Productos")
lbl_productos.pack(anchor="w", pady=(12, 0), padx=12)

tree_productos = ttk.Treeview(
    tables_frame,
    columns=("Código Producto", "Nombre Producto", "Cantidad", "Precio"),
    show="headings",
    height=8,
)
for col in ("Código Producto", "Nombre Producto", "Cantidad", "Precio"):
    tree_productos.heading(col, text=col)
    tree_productos.column(col, anchor="center", width=170)
tree_productos.pack(fill="x", padx=12, pady=8)

# --- ESTILO DE TABLAS ---
style = ttk.Style()
style.theme_use("clam")
style.configure(
    "Treeview",
    background="#22252A",
    fieldbackground="#22252A",
    foreground="#FFFFFF",
    rowheight=28,
    borderwidth=0,
    font=("Segoe UI", 11),
)
style.configure(
    "Treeview.Heading",
    background="#2E3238",
    foreground="#FFA726",
    font=("Segoe UI", 12, "bold"),
    borderwidth=0,
)
style.map("Treeview", background=[("selected", PRIMARY_HOVER)])

# --- BOTÓN INICIAR ---

def _run_iniciar_proceso_thread():
    try:
        root.after(0, lambda: btn_iniciar.configure(state="disabled"))
        root.after(0, lambda: mostrar_toast("Proceso iniciado en background", tipo="info", titulo="Proceso"))
        iniciar_proceso()
        root.after(0, lambda: btn_iniciar.configure(state="normal"))
        root.after(0, lambda: mostrar_toast("Proceso completado ✅", tipo="success", titulo="Completado"))
        # Animación de éxito rápida
        root.after(0, lambda: quick_pulse_animation(btn_iniciar, "#4CAF50"))
    except Exception as e:
        root.after(0, lambda: mostrar_toast(f"Error en proceso: {e}", tipo="error", titulo="Error"))
        root.after(0, lambda: btn_iniciar.configure(state="normal"))
        # Animación de error rápida
        root.after(0, lambda: quick_pulse_animation(btn_iniciar, "#F44336"))

def _on_iniciar_proceso():
    mostrar_toast("Iniciando proceso...", tipo="info", titulo="Iniciar")
    quick_pulse_animation(btn_iniciar, "#FFD54F")
    t = threading.Thread(target=_run_iniciar_proceso_thread, daemon=True)
    t.start()

btn_iniciar = ctk.CTkButton(
    main_frame,
    text="🚀 Iniciar proceso",
    command=_on_iniciar_proceso,
    fg_color=PRIMARY,
    hover_color=PRIMARY_HOVER,
    text_color="#FFFFFF",
    font=("Segoe UI", 15, "bold"),
    corner_radius=24,
    width=250,
    height=48,
    state="disabled",
)
btn_iniciar.pack(pady=16)

btn_iniciar.bind("<Enter>", lambda e: simple_button_hover(btn_iniciar, True))
btn_iniciar.bind("<Leave>", lambda e: simple_button_hover(btn_iniciar, False))

# --- FOOTER ---
footer = ctk.CTkFrame(pantalla_facturas, fg_color=CARD_BG, corner_radius=0)
footer.pack(fill="x", side="bottom")
lbl_footer = ctk.CTkLabel(footer, text="Desarrollado por Area de desarrollo de Healthy", font=("Segoe UI", 11), text_color=TEXT_SECOND)
lbl_footer.pack(pady=8)

# =========================
# PANTALLA DE INVENTARIO
# =========================
pantalla_inventario = ctk.CTkFrame(contenedor, fg_color=BG)
pantallas["inventario"] = pantalla_inventario

# === HEADER SUPERIOR ===
header_inventario = ctk.CTkFrame(pantalla_inventario, fg_color=CARD_BG, corner_radius=20)
header_inventario.pack(fill="x", padx=20, pady=12)

lbl_titulo_inventario = ctk.CTkLabel(
    header_inventario,
    text="Sistema de Inventario Automatizado",
    font=("Segoe UI", 24, "bold"),
    text_color=TEXT_MAIN,
)
lbl_titulo_inventario.pack(side="left", padx=20, pady=12)

# Botón volver
def _on_volver_inventario():
    mostrar_toast("Volviendo al menú", tipo="info", titulo="Volver")
    mostrar_pantalla("menu")

btn_volver_inventario = ctk.CTkButton(
    header_inventario,
    text="↩ Volver",
    command=_on_volver_inventario,
    width=90,
    height=36,
    corner_radius=16,
    fg_color="#5EA0FF",
    hover_color="#7AB8FF",
    font=("Segoe UI", 13, "bold"),
)
btn_volver_inventario.pack(side="right", padx=10, pady=10)

btn_volver_inventario.bind("<Enter>", lambda e: simple_button_hover(btn_volver_inventario, True))
btn_volver_inventario.bind("<Leave>", lambda e: simple_button_hover(btn_volver_inventario, False))

# Botón cerrar
def cerrar_app_inventario():
    if confirmar_salida("Salir", "¿Deseas cerrar la aplicación?"):
        mostrar_toast("Cerrando aplicación...", tipo="info", titulo="Cerrando")
        root.after(300, root.destroy)
    else:
        mostrar_toast("Salida cancelada", tipo="info", titulo="Cancelado")

btn_close_inventario = ctk.CTkButton(
    header_inventario,
    text="✕",
    command=cerrar_app_inventario,
    width=40,
    height=40,
    corner_radius=20,
    fg_color="#FF5252",
    hover_color="#FF6B6B",
    font=("Segoe UI", 14, "bold"),
)
btn_close_inventario.pack(side="right", padx=10, pady=10)

btn_close_inventario.bind("<Enter>", lambda e: simple_button_hover(btn_close_inventario, True))
btn_close_inventario.bind("<Leave>", lambda e: simple_button_hover(btn_close_inventario, False))

# === CUERPO PRINCIPAL ===
main_frame_inventario = ctk.CTkScrollableFrame(pantalla_inventario, fg_color=BG, corner_radius=0)
main_frame_inventario.pack(fill="both", expand=True, padx=30, pady=10)

# --- BOTONES DE PROCESO ---
btn_frame_inventario = ctk.CTkFrame(main_frame_inventario, fg_color=CARD_BG, corner_radius=25)
btn_frame_inventario.pack(pady=18)

def _on_descargar_informes():
    mostrar_toast("Iniciando descarga de informes...", tipo="info", titulo="Descarga")
    quick_pulse_animation(btn_descargar_informes, "#FFD54F")
    try:
        descargar_informes_inventario()
    except Exception as e:
        mostrar_toast(f"Error al descargar informes: {e}", tipo="error", titulo="Error")

btn_descargar_informes = ctk.CTkButton(
    btn_frame_inventario,
    text="📥 Descargar Informes",
    command=pedir_fecha_informes,
    corner_radius=25,
    fg_color=ACCENT,
    hover_color=ACCENT_HOVER,
    text_color="#000000",
    font=("Segoe UI", 14, "bold"),
    width=220,
)
btn_descargar_informes.pack(side="left", padx=12, pady=12)

def _on_procesar_informes():
    mostrar_toast("Procesando informes descargados...", tipo="info", titulo="Procesamiento")
    quick_pulse_animation(btn_procesar_informes, "#4FC3F7")
    try:
        procesar_informes_inventario()
    except Exception as e:
        mostrar_toast(f"Error al procesar informes: {e}", tipo="error", titulo="Error")

btn_procesar_informes = ctk.CTkButton(
    btn_frame_inventario,
    text="🔄 Procesar Informes",
    command=_on_procesar_informes,
    corner_radius=25,
    fg_color=PRIMARY,
    hover_color=PRIMARY_HOVER,
    text_color="#FFFFFF",
    font=("Segoe UI", 14, "bold"),
    width=220,
    state="disabled",
)
btn_procesar_informes.pack(side="left", padx=12, pady=12)
# Añadir hover simple
btn_descargar_informes.bind("<Enter>", lambda e: simple_button_hover(btn_descargar_informes, True))
btn_descargar_informes.bind("<Leave>", lambda e: simple_button_hover(btn_descargar_informes, False))
btn_procesar_informes.bind("<Enter>", lambda e: simple_button_hover(btn_procesar_informes, True))
btn_procesar_informes.bind("<Leave>", lambda e: simple_button_hover(btn_procesar_informes, False))

# --- INFORMACIÓN DE ARCHIVO Y ESTADO ---
info_frame_inventario = ctk.CTkFrame(main_frame_inventario, fg_color=BG)
info_frame_inventario.pack(pady=12, fill="x")

lbl_info_inventario = ctk.CTkLabel(
    info_frame_inventario,
    text="Esperando descarga de informes...",
    font=("Segoe UI", 14, "bold"),
    text_color=TEXT_SECOND,
)
lbl_info_inventario.pack(anchor="w", pady=(0, 4))

lbl_estado_inventario = ctk.CTkLabel(
    info_frame_inventario,
    text="Estado: Listo para comenzar",
    font=("Segoe UI", 12),
    text_color=TEXT_SECOND,
)
lbl_estado_inventario.pack(anchor="w")

# === SECCIÓN INFORMATIVA ===
info_frame = ctk.CTkFrame(main_frame_inventario, fg_color=CARD_BG, corner_radius=20)
info_frame.pack(fill="x", padx=10, pady=10)

lbl_info_titulo = ctk.CTkLabel(
    info_frame,
    text="📋 Proceso de Inventario Automatizado",
    font=("Segoe UI", 16, "bold"),
    text_color=ACCENT,
    anchor="w",
)
lbl_info_titulo.pack(anchor="w", pady=(10, 5), padx=12)

lbl_info_desc = ctk.CTkLabel(
    info_frame,
    text="Este proceso automatizado:\n• Descarga 9 informes de inventario (uno por cada clínica)\n• Escribe automáticamente los códigos de cada clínica\n• Llena los campos necesarios y descarga en formato XLSX\n• Combina todos los datos en un solo archivo Excel\n• Guarda el resultado como 'Inventario_Combinado_{Fecha_Hora}'",
    font=("Segoe UI", 12),
    text_color=TEXT_SECOND,
    anchor="w",
    justify="left",
)
lbl_info_desc.pack(anchor="w", pady=(0, 10), padx=12)

# --- FOOTER ---
footer_inventario = ctk.CTkFrame(pantalla_inventario, fg_color=CARD_BG, corner_radius=0)
footer_inventario.pack(fill="x", side="bottom")
lbl_footer_inventario = ctk.CTkLabel(footer_inventario, text="Desarrollado por Area de desarrollo de Healthy", font=("Segoe UI", 11), text_color=TEXT_SECOND)
lbl_footer_inventario.pack(pady=8)

# MOSTRAR PANTALLA INICIAL - Ahora es login
mostrar_pantalla("login")

# Solo iniciamos mainloop si no hay otro en el archivo
if not hasattr(root, "_DataSpectra_mainloop_started"):
    root._DataSpectra_mainloop_started = True
    root.mainloop()